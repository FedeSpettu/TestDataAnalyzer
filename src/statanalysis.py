from tkinter import filedialog
import tkinter as tk
import numpy as np
import pandas as pd
from datetime import datetime
import openpyxl
import csv
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from src.global_var import *
#from src.loading import *
from .usefull_functions import auto_detect_delimiter, is_date_column, remove_special_characters, is_date_column2, remove_special_characters_from_list
from tkinter import messagebox as mbox
import sys
from tkinter import Tk, Checkbutton, Button, Scrollbar, messagebox
import customtkinter as ctk
import os.path
from openpyxl.drawing.image import Image as ExcelImage
from src import dataload as dl
import seaborn as sns
import matplotlib.dates as mdates
import re
import pandas as pd
from datetime import datetime, timedelta
from src.interactive_tkinter_called import launch_interactive_plot


def reset_gui2( elements):
   
    global plot_dictionary
    plot_dictionary= {
    }
    global selection 
    selection = {
        'File1': [],
        'File2': []
    }
    global folder_path
    folder_path=[]
    global output_path
    output_path=''
    global output_file
    output_file=''
    global k
    k=-1
    global z
    z=0
    global j
    j=0
    global currentpage
    currentpage=0
    global min_data_rows
    min_data_rows = 2 
    
    if os.path.isfile('options.txt'):
        os.remove('options.txt')

    if os.path.exists('options1.txt'):
        os.remove('options1.txt')
    
    if os.path.exists('options2.txt'):
        os.remove('options2.txt')

    if os.path.isfile('output0.csv'):
        os.remove('output0.csv')
    if os.path.isfile('output1.csv'):
        os.remove('output1.csv')
    if os.path.isfile('options.txt'):
        os.remove('options.txt')
    if os.path.isfile('check.csv'):
        os.remove('check.csv')
    if os.path.isfile('data.csv'):
        os.remove('data.csv')
    if os.path.isfile('backup.csv'):
        os.remove('backup.csv')
    if os.path.isfile('backupdf2.csv'):
        os.remove('backupdf2.csv')
    if os.path.isfile('backupprova.csv'):    
        os.remove('backupprova.csv')
    if os.path.isfile('prova2.csv'):
        os.remove('prova2.csv')
   
    if os.path.isfile('plot.png'):    
        os.remove('plot.png')
    if os.path.isfile('prova.csv'):    
        os.remove('prova.csv')
    if os.path.isfile('prova1.csv'):    
        os.remove('prova1.csv')
    if os.path.isfile('prova3.csv'):    
        os.remove('prova3.csv')

    if os.path.isfile('backupdf1.csv'):    
        os.remove('backupdf1.csv')
    if os.path.isfile('backupdoasjfouieqhfiufho.csv'):    
        os.remove('backupdoasjfouieqhfiufho.csv')
    if os.path.isfile('output0_modified.csv'):
        os.remove('output0_modified.csv')
    if os.path.isfile('output1_modified.csv'):
        os.remove('output1_modified.csv')
    if os.path.isfile('output0check.csv'):    
        os.remove('output0check.csv')
    if os.path.isfile('output1check.csv'):    
        os.remove('output1check.csv')
    if os.path.isfile('options_event.txt'):
        os.remove('options_event.txt')
    for element in elements:
        
        if isinstance(element, tk.Listbox):
            element.delete(0, tk.END)
        

def truncate_text(text, max_length=30):
    """Returns the text truncated to max_length characters (including an ellipsis) if needed."""
    if len(text) > max_length:
        return text[:max_length - 3] + "..."
    return text

def select_output(folder_label3, clean_paths,
        output_folder_label,
        file1_optionmenu,
        file1_optionmenu_var,
        file2_optionmenu_var,
        file1_listboxfile,
        limit1_entry,
        limit2_entry, 
        nvalues_entry,
        checkbox_threshold,
        checkbox_align,
        unit_var,
        checkbox_plot,
        condition_var,
        pass_fail_var,
        file1_column_option,
        file1_listbox,
        output_frame,
        file1_frame,
        file2_frame,
        analysis_frame,
        checkbox_event,
        start_event_var,
        end_event_var,
        file1_folder_label,
        file2_folder_label,
        start_event_option,
        end_event_option,
        event_listbox2,
        event_listbox1,
        event_filter_entry1,
        event_filter_entry2,
        unit_option_menu,
        file2_column_option_var,
        file1_column_option_var,
        checkbox_advance):
    elements = [
        clean_paths,
        output_folder_label,
        file1_optionmenu,
        file1_optionmenu_var,
        file2_optionmenu_var,
        file1_listboxfile,
        limit1_entry,
        limit2_entry, 
        nvalues_entry,
        checkbox_threshold,
        checkbox_align,
        unit_var,
        checkbox_plot,
        condition_var,
        pass_fail_var,
        file1_column_option,
        file1_listbox,
        output_frame,
        file1_frame,
        file2_frame,
        analysis_frame,
        checkbox_event.get(),
        start_event_var,
        end_event_var,
        file1_optionmenu,
        file1_folder_label,
        file2_folder_label,
        file1_listbox,
        file1_listboxfile,
        checkbox_event,
        start_event_option,
        end_event_option,
        event_listbox2,
        event_listbox1,
        event_filter_entry1,
        event_filter_entry2,
        start_event_var,
        end_event_var,
        file1_listbox,
        file1_optionmenu,  # OptionMenu widget
        limit1_entry,
        limit2_entry,
        nvalues_entry,
        unit_option_menu,  
        file2_column_option_var,
        file1_column_option_var,
        checkbox_advance
    ]
    try:
        global output_path
        global output_file
        
        output_path = filedialog.askdirectory()
        if len(output_path)<2:
            return
        output_path=output_path + '\\'
        if output_path:
            folder_label3.configure(text=truncate_text(output_path, 200))
        output_file=output_path + 'Result' + str(datetime.now().strftime("%Y-%m-%d %H-%M-%S")+'.xlsx')
        wb = openpyxl.Workbook()
        sheet_name = "Files"  # Change this to your desired sheet name
        
        header = ['Folder1', 'Folder2', 'Outcome']

        ws = wb.active
        ws.title = sheet_name
        ws.append(header)
        wb.save(output_file)
        
    except Exception as e:
        root=tk.Tk()
        root.withdraw()
        print(e)
        error_message = e.args
        messagebox.showerror("Critical Error", str(error_message), parent=root)

def convert_to_relative_time(ms_col, reference_time):
    try:
    # Convert column to pandas Series
        ms_series = pd.Series(ms_col)

        # Get reference time as timedelta
        ref_td = pd.to_timedelta(reference_time)

        # Convert column to timedelta and add reference time
        relative_td = ref_td + pd.to_timedelta(ms_series, unit='ms')

        # Convert timedelta to string 
        relative_strings = relative_td.astype(str)

        return relative_strings
    except Exception as e:
        root=tk.Tk()

        root.withdraw()
        print(e)
        error_message = e.args
        messagebox.showerror("Critical Error", str(error_message))

def apply_formulas_to_column(df,reftime,column_date):
    # Validate input column is numeric
    try:
        
        time_column = df[column_date]

            # Make sure column is numeric before time conversion 
        df = df.copy()
        time_column = pd.to_numeric(time_column, errors='coerce')
    
        reftime_str = reftime.strftime('%H:%M:%S')
        reference_timedelta = pd.to_timedelta(reftime_str)
        
        if not isinstance(time_column[0], (int, float, str)):
            relative_times = convert_to_relative_time(df[column_date], reference_timedelta)
        
            df[column_date] = relative_times.apply(lambda x: str(x).split()[-1])

            return df
            #raise TypeError("Input column must be numeric or string")
        # Add reference time to column
        time_column = pd.to_timedelta(time_column, unit='s') + reference_timedelta

        # Round to seconds and convert to string
        time_column = time_column.dt.round('1s').apply(lambda x: str(x).split()[-1])
        # Replace column in copied DataFrame
        df[column_date] = time_column

        return df
    except Exception as e:
        root=tk.Tk()
        root.withdraw()
        print(e)
        error_message = e.args
        messagebox.showerror("Critical Error", str(error_message))

# Main analysis function

import threading
import queue
from PIL import Image, ImageTk, ImageSequence

class LoadingScreen_mult:
    def __init__(self, root, minimum, maximum, threshold, checkbox, checkbox1,var_unit,checkbox2,text_input4, text_input5,clickedfolder2,clickedfolder1, var_unit3, var_unit2, i, n_files, clean_paths, drop1, drop2, file_list2, checkbox_var3, clickedeventstart,clickedeventend, enable_plot):
        self.root = root
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        self.root.title("Loading columns")
        self.root.geometry(f"+{screen_width // 2 - root.winfo_width() // 2}+{screen_height // 2 - root.winfo_height() // 2}")
        #self.root.geometry("300x330+400+300")  # Set your desired size and position

        self.after_queue = queue.Queue()
        gif_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '', 'giphy.gif')
        self.loading_gif = Image.open(gif_path)
        self.frames = [ImageTk.PhotoImage(frame.copy()) for frame in ImageSequence.Iterator(self.loading_gif)]

        self.loading_label = tk.Label(root)
        self.loading_label.pack()

        self.frame_num = 0
        self.loading = True
        self.load_thread = threading.Thread(target=self.load_big_file_mult, args=(root, minimum, maximum, threshold, checkbox, checkbox1,var_unit,checkbox2,text_input4, text_input5,clickedfolder2,clickedfolder1, var_unit3, var_unit2, i, n_files, clean_paths, drop1, drop2, file_list2, checkbox_var3, clickedeventstart,clickedeventend, enable_plot))
        self.load_thread.start()

        self.update_frames_mult()

    def update_frames_mult(self):
        if self.loading_label.winfo_exists():
            self.loading_label.configure(image=self.frames[self.frame_num])
            self.loading_label.image = self.frames[self.frame_num]
            self.loading_label.update()  # Force update

            # Move to the next frame
            self.frame_num = (self.frame_num + 1) % len(self.frames)

        if self.loading:
            self.root.after(100, self.update_frames_mult)  # Adjust the delay as needed

    def load_big_file_mult(self,root, minimum, maximum, threshold, checkbox, checkbox1,var_unit,checkbox2,text_input4, text_input5,clickedfolder2,clickedfolder1, var_unit3, var_unit2, i, n_files, clean_paths, drop1, drop2, file_list2, checkbox_var3, clickedeventstart,clickedeventend, enable_plot):
        i=0
        try:
            while self.loading:
                # Simulate loading a big file (replace this with your actual file loading logic)
                

                path=clean_paths[i]
                
                delimiter=auto_detect_delimiter(path)
                df = pd.read_csv(path, sep=delimiter, encoding='latin1')
                indices = []
                for file, headers in selection.items():
                    for header in headers:
                        if header in df.columns:
                            idx = df.columns.get_loc(header)
                            indices.append(idx)
                
                selection['File1']= df.columns[indices]
                filename = os.path.basename(path)
                clickedfolder1.set(filename)
                dl.load_data_mult(path)
                analyze_files(minimum, maximum, threshold, checkbox, checkbox1,var_unit,checkbox2,text_input4, text_input5,clickedfolder2,clickedfolder1, var_unit3, var_unit2, self.loading_label, root, i, n_files, checkbox_var3, clickedeventstart,clickedeventend, enable_plot)
                i=i+1
                if i==n_files:
                    clean_paths=[]
                    self.loading = False
                    self.loading_label.destroy()
                    root.withdraw()
                    # if os.path.isfile('options_multi.txt'):
                    #     os.remove('options_multi.txt')
                    # if os.path.isfile('options.txt'):
                    #     os.remove('options.txt')
                    # if os.path.isfile('options1.txt'):
                    #     os.remove('options1.txt')
                    # if os.path.isfile('options2.txt'):    
                    #     os.remove('options2.txt')
                    # if os.path.isfile('output0.csv'):
                    #     os.remove('output0.csv')
                    # if os.path.isfile('output1.csv'):
                    #     os.remove('output1.csv')
                    # if os.path.isfile('options.txt'):
                    #     os.remove('options.txt')
                    if os.path.isfile('check.csv'):
                        os.remove('check.csv')
                    if os.path.isfile('data.csv'):
                        os.remove('data.csv')
                    if os.path.isfile('backup.csv'):
                        os.remove('backup.csv')
                    if os.path.isfile('backupdf2.csv'):
                        os.remove('backupdf2.csv')
                    if os.path.isfile('backupprova.csv'):    
                        os.remove('backupprova.csv')
                    if os.path.isfile('prova2.csv'):
                        os.remove('prova2.csv')
                    
                    if os.path.isfile('plot.png'):    
                        os.remove('plot.png')
                    if os.path.isfile('prova.csv'):    
                        os.remove('prova.csv')
                    if os.path.isfile('prova1.csv'):    
                        os.remove('prova1.csv')
                    if os.path.isfile('prova3.csv'):    
                        os.remove('prova3.csv')
                 
                    if os.path.isfile('backupdf1.csv'):    
                        os.remove('backupdf1.csv')
                    if os.path.isfile('backupdoasjfouieqhfiufho.csv'):    
                        os.remove('backupdoasjfouieqhfiufho.csv')
                    if os.path.isfile('output0_modified.csv'):
                        os.remove('output0_modified.csv')
                    if os.path.isfile('output1_modified.csv'):
                        os.remove('output1_modified.csv')
                    if os.path.isfile('output0check.csv'):    
                        os.remove('output0check.csv')
                    if os.path.isfile('output1check.csv'):    
                        os.remove('output1check.csv')
                    #if os.path.isfile('options_event.txt'):
                    #    os.remove('options_event.txt')
                    messagebox.showinfo("Done", "The analysis have been saved", parent=root)  
                #time.sleep(0.1)  # Simulate a short delay for demonstration
            
        except Exception as e:
        
            print(e)
            error_message = e.args
            messagebox.showerror("Critical Error", str(error_message))
            
          

def loading_fun_mult(minimum, maximum, threshold, checkbox, checkbox1,var_unit,checkbox2,text_input4, text_input5,clickedfolder2,clickedfolder1, var_unit3, var_unit2, i, n_files, clean_paths, drop1, drop2, file_list2, frame1, frame2, frame3, input_frame , checkbox_var3, clickedeventstart,clickedeventend, enable_plot):
    root_load = tk.Toplevel()
    root_load.overrideredirect(True)
    screen_width = root_load.winfo_screenwidth()
    screen_height = root_load.winfo_screenheight()

    root_load.geometry(f"+{screen_width // 2 - 165}+{screen_height // 2 - 165}")
    app = LoadingScreen_mult(root_load, minimum, maximum, threshold, checkbox, checkbox1,var_unit,checkbox2,text_input4, text_input5,clickedfolder2,clickedfolder1, var_unit3, var_unit2, i, n_files, clean_paths, drop1, drop2, file_list2, checkbox_var3, clickedeventstart,clickedeventend, enable_plot)
    root_load.mainloop()

class LoadingScreen:
    def __init__(self, root, minimum, maximum, threshold, checkbox, checkbox1,var_unit,checkbox2,text_input4, text_input5,clickedfolder2,clickedfolder1, var_unit3, var_unit2, i, n_files,  checkbox_var3, clickedeventstart,clickedeventend, enable_plot):
        self.root = root
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        self.root.title("Loading columns")
        self.root.geometry(f"+{screen_width // 2 - root.winfo_width() // 2}+{screen_height // 2 - root.winfo_height() // 2}")
        #self.root.geometry("300x330+400+300")  # Set your desired size and position

        self.after_queue = queue.Queue()
        gif_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '', 'giphy.gif')
        self.loading_gif = Image.open(gif_path)
        self.frames = [ImageTk.PhotoImage(frame.copy()) for frame in ImageSequence.Iterator(self.loading_gif)]

        self.loading_label = tk.Label(root)
        self.loading_label.pack()

        self.frame_num = 0
        self.loading = True
        self.load_thread = threading.Thread(target=self.load_big_file, args=(root, minimum, maximum, threshold, checkbox, checkbox1,var_unit,checkbox2,text_input4, text_input5,clickedfolder2,clickedfolder1, var_unit3, var_unit2, i, n_files , checkbox_var3, clickedeventstart,clickedeventend, enable_plot))
        self.load_thread.start()

        self.update_frames()

    def update_frames(self):
        if self.loading_label.winfo_exists():
            self.loading_label.configure(image=self.frames[self.frame_num])
            self.loading_label.image = self.frames[self.frame_num]
            self.loading_label.update()  # Force update

            # Move to the next frame
            self.frame_num = (self.frame_num + 1) % len(self.frames)

        if self.loading:
            self.root.after(100, self.update_frames)  # Adjust the delay as needed

    def load_big_file(self,root, minimum, maximum, threshold, checkbox, checkbox1,var_unit,checkbox2,text_input4, text_input5,clickedfolder2,clickedfolder1, var_unit3, var_unit2, i, n_files, checkbox_var3, clickedeventstart,clickedeventend, enable_plot):
        try:
            while self.loading:
                # Simulate loading a big file (replace this with your actual file loading logic)
                analyze_files(minimum, maximum, threshold, checkbox, checkbox1,var_unit,checkbox2,text_input4, text_input5,clickedfolder2,clickedfolder1, var_unit3, var_unit2, self.loading_label, root, i, n_files, checkbox_var3, clickedeventstart,clickedeventend, enable_plot)
                self.loading = False
                
                #time.sleep(0.1)  # Simulate a short delay for demonstration
        except Exception as e:
        
            print(e)
            error_message = e.args
            dialog = (root, error_message, "Critical Error")
            dialog.mainloop()
          

def loading_fun(minimum, maximum, threshold, checkbox, checkbox1,var_unit,checkbox2,text_input4, text_input5,clickedfolder2,clickedfolder1, var_unit3, var_unit2, i, n_files, frame1, frame2, frame3, input_frame, checkbox_var3, clickedeventstart,clickedeventend, enable_plot):
    root_load = tk.Toplevel()
    root_load.overrideredirect(True)
    screen_width = root_load.winfo_screenwidth()
    screen_height = root_load.winfo_screenheight()
    root_load.geometry(f"+{screen_width // 2 - 165}+{screen_height // 2 - 165}")
    
     
    app = LoadingScreen(root_load, minimum, maximum, threshold, checkbox, checkbox1,var_unit,checkbox2,text_input4, text_input5,clickedfolder2,clickedfolder1, var_unit3, var_unit2, i, n_files, checkbox_var3, clickedeventstart,clickedeventend, enable_plot)
    root_load.mainloop()

def remove_spaces_and_replace_with_comma(input_file_path):
    try:
        file_name, file_extension = os.path.splitext(input_file_path)
        output_file_path = f"{file_name}_modified.csv"
        additional_output_file_path = 'your_additional_output_file.csv'

        with open(input_file_path, 'r', newline='') as infile:
            with open(output_file_path, 'w', newline='') as outfile, open(additional_output_file_path, 'w', newline='') as additional_outfile:
                for line in infile:
                    # Use a regular expression to replace spaces between letters and numbers with a comma
                    line_modified = re.sub(r'([a-zA-Z])\s+([\d.-])', r'\1,\2', line)

                    # Remove any remaining spaces between non-space characters
                    line_modified = re.sub(r'(?<=[^\s])\s+(?=[^\s])', ',', line_modified)

                    outfile.write(line_modified)
                    additional_outfile.write(line_modified)
        # file_name, file_extension = os.path.splitext(input_file_path)
        # output_file_path = f"{file_name}_modified.csv"
        # additional_output_file_path = 'C:\src\your_additional_output_file.csv'
        # with open(input_file_path, 'r', newline='') as infile:
        #     with open(output_file_path, 'w', newline='') as outfile, open(additional_output_file_path, 'w', newline='') as additional_outfile:
        #         for line in infile:
        #             # Remove spaces and replace with a comma
        #             line = line.replace(' ', ',')
        #             outfile.write(line)
        #             additional_outfile.write(line)
    except Exception as e:
        
        root=tk.Tk()
        root.withdraw()
        print(e)
        error_message = e.args
        messagebox.showerror("Critical Error", str(error_message))

def is_numerical(value):
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False

def insert_seconds_progressively(time_column, seconds_to_start, seconds_to_add):


  time_column = pd.to_datetime(time_column)
  last_minute = time_column.iloc[0].minute

  for i in range(len(time_column)):
    if time_column.iloc[i].minute != last_minute:
      seconds_to_start = 0
      last_minute = time_column.iloc[i].minute

    time_column.iloc[i] += pd.Timedelta(seconds=seconds_to_start)
    seconds_to_start += seconds_to_add

  return time_column


def align_dataframes(df1, df2, time_column1, time_column2):
    import pandas as pd

def align_dataframes(df1, df2, time_column1, time_column2):
    """
    Allinea solo l'inizio e la fine dei due DataFrame ad un tempo comune (global_start e global_end)
    aggiungendo delle righe di zero padding (a cadenza di 1 secondo) solo per coprire le parti mancanti.
    
    Ad esempio, se per df1:
      - start1 > global_start: viene creato un padding all'inizio (da global_start fino a start1 - 1 sec)
      - end1 < global_end:  viene creato un padding finale (da end1 + 1 sec fino a global_end)
    
    Lo stesso per df2.  
    I due DataFrame mantengono le loro colonne temporali indipendenti, e i nuovi timestamp
    sono formattati come hh:mm:ss.
    
    Parameters:
      df1, df2          : DataFrame da allineare
      time_column1      : nome della colonna temporale in df1
      time_column2      : nome della colonna temporale in df2
      
    Returns:
      tuple: (df1_allineato, df2_allineato)
    """
    # Creiamo copie dei DataFrame originali
    df1 = df1.copy()
    df2 = df2.copy()
    
    # Converte le colonne temporali in datetime
    df1[time_column1] = pd.to_datetime(df1[time_column1])
    df2[time_column2] = pd.to_datetime(df2[time_column2])
    
    # Calcola gli estremi (start ed end) per ciascun DataFrame
    start1 = df1[time_column1].min()
    end1   = df1[time_column1].max()
    start2 = df2[time_column2].min()
    end2   = df2[time_column2].max()
    
    # Determina il global start ed end (minimo dei due start, massimo dei due end)
    global_start = min(start1, start2)
    global_end   = max(end1, end2)
    delta_sec=(start1 - start2).total_seconds()

    # Identifica le colonne dei valori (cioè tutte quelle tranne la colonna temporale)
    value_columns1 = [col for col in df1.columns if col != time_column1]
    value_columns2 = [col for col in df2.columns if col != time_column2]
    
    # ----- Padding per df1 -----
    df1_list = []
    
    # Se il primo timestamp di df1 è successivo a global_start, crea il padding iniziale
    if start1 > global_start:
        pad_index = pd.date_range(start=global_start, end=start1 - pd.Timedelta(seconds=1), freq='s')
        pad_df1 = pd.DataFrame({time_column1: pad_index})
        for col in value_columns1:
            pad_df1[col] = 0
        df1_list.append(pad_df1)
    
    # Aggiungi il DataFrame originale
    df1_list.append(df1)
    
    # Se l'ultimo timestamp di df1 è precedente a global_end, crea il padding finale
    if end1 < global_end:
        pad_index = pd.date_range(start=end1 + pd.Timedelta(seconds=1), end=global_end, freq='s')
        pad_df1_end = pd.DataFrame({time_column1: pad_index})
        for col in value_columns1:
            pad_df1_end[col] = 0
        df1_list.append(pad_df1_end)
    
    # Concatena le parti e riordina in base alla colonna temporale
    df1_aligned = pd.concat(df1_list, ignore_index=True)
    df1_aligned = df1_aligned.sort_values(by=time_column1).reset_index(drop=True)
    
    # ----- Padding per df2 -----
    df2_list = []
    
    if start2 > global_start:
        pad_index = pd.date_range(start=global_start, end=start2 - pd.Timedelta(seconds=1), freq='s')
        pad_df2 = pd.DataFrame({time_column2: pad_index})
        for col in value_columns2:
            pad_df2[col] = 0
        df2_list.append(pad_df2)
    
    df2_list.append(df2)
    
    if end2 < global_end:
        pad_index = pd.date_range(start=end2 + pd.Timedelta(seconds=1), end=global_end, freq='s')
        pad_df2_end = pd.DataFrame({time_column2: pad_index})
        for col in value_columns2:
            pad_df2_end[col] = 0
        df2_list.append(pad_df2_end)
    
    df2_aligned = pd.concat(df2_list, ignore_index=True)
    df2_aligned = df2_aligned.sort_values(by=time_column2).reset_index(drop=True)
    
    # Formatto le colonne dei tempi in hh:mm:ss
    df1_aligned[time_column1] = df1_aligned[time_column1].dt.strftime('%H:%M:%S')
    df2_aligned[time_column2] = df2_aligned[time_column2].dt.strftime('%H:%M:%S')
    
    return df1_aligned, df2_aligned, delta_sec




def analyze_files(minimum, maximum, threshold, checkbox, checkbox1,var_unit,checkbox2,text_input4, text_input5,clickedfolder2,clickedfolder1, var_unit3, var_unit2, loading_label, root,itera_tion, n_files, checkbox_var3, clickedeventstart,clickedeventend, enable_plot):
    global output_path
    global currentpage
    global output_file
    try:
        if os.path.isfile('options1.txt'):
            lines = []
            with open('options1.txt') as f:
                for line in f:
                    lines.append(line.strip())
            selection['File1'] = lines
        if os.path.isfile('options2.txt'):
            lines = []
            with open('options2.txt') as f:
                for line in f:
                    lines.append(line.strip())
            selection['File2'] = lines
        temp_sel1=selection['File1']
        temp_sel2=selection['File2']
        if checkbox_var3:
            temp_sel1.append('Event')
        print(temp_sel1)
        select=0
        count=0
        count1=0
        print(clickedfolder2.get(), clickedfolder1.get())

        if clickedfolder2.get() != 'Select File' and clickedfolder1.get()=='Select File':
            files_folder1= ['']
            files_folder2=[clickedfolder2.get()]
        elif clickedfolder1.get() != 'Select File'and clickedfolder2.get()=='Select File':
            files_folder2= ['']
            files_folder1=[clickedfolder1.get()]
        else:
            files_folder1= [clickedfolder1.get()]
            files_folder2=[clickedfolder2.get()]
            select=2

        if checkbox1 and select==2: #if alignment is true and select is 2 then it will run the code below filling the cells with the files in the folder
            wb = load_workbook(output_file)
            ws = wb['Files']
            start_row = ws.max_row + 1
            for item in files_folder1:
                if item:
                    cell = ws.cell(row=start_row, column=1, value=item)
                if files_folder2:
                    cell = ws.cell(row=start_row, column=2, value=str(files_folder2))
                start_row += 1
            wb.save(output_file)

            delimiter1 = auto_detect_delimiter('output0.csv') #auto detect delimiter of the first file
            delimiter2 = auto_detect_delimiter('output1.csv') #auto detect delimiter of the second file
            df1 = pd.read_csv('output0.csv', sep=delimiter1, encoding='UTF-8', engine='python')
            df2 = pd.read_csv('output1.csv', sep=delimiter2, encoding='UTF-8', engine='python')
            df1=remove_special_characters(df1)
            df2=remove_special_characters(df2)
            try: 
                df1.columns = df1.columns.str.replace('[', '').str.replace(']', '') #remove special characters from the column names

                if checkbox_var3: #if checkbox_var3 is true then it will run the code below taking the event column and adding it to the selection
                    df1 = df1[temp_sel1]
                else:
                    df1= df1[selection['File1']]
                df1 = df1.apply(lambda col: col.map(lambda x: x.replace('[', '').replace(']', '') if isinstance(x, str) else x))
                timecol=df1.columns[0]

                for col in df1.columns:
                    if col != 'Event' and col!= timecol: ##if the column is not the event column then it will run the code below

                        for i in range(len(df1[col])):
                            try:
                                # Try to convert the individual value to numeric, handling errors and replacing commas with dots
                                df1.at[i, col] = pd.to_numeric(str(df1.at[i, col]).replace(',', '.'), errors='coerce')

                                # Check if the value is non-numeric
                                if not pd.api.types.is_numeric_dtype(df1.at[i, col]):
                                    # If the value is non-numeric, replace it with NaN
                                    df1.at[i, col] = np.nan
                            except:
                                # Use a different approach to handle non-string values
                                df1.at[i,col] = float(df1.at[i, col]) if pd.api.types.is_numeric_dtype(df1.at[i, col]) else np.nan
                    else:
                        pass
                df1.to_csv('output0.csv', index=False)
                
            except Exception as e:
            
                remove_spaces_and_replace_with_comma('output0.csv')
                file_name, file_extension = os.path.splitext('output0.csv')
                output_file_path = f"{file_name}_modified.csv"
                with open(output_file_path, 'r') as f:
                    reader = csv.reader(f, delimiter=',')
                    headers = next(reader)  # Read the first row (header)

                    # If the header row is empty, read the next non-empty row as the header
                    while not any(headers):
                        
                        headers = next(reader)
                for i, element in enumerate(headers):
                    headers[i] = element.replace('[', '').replace(']','')
                
                df1.to_csv('output0.csv', index=False)
                df1 = pd.read_csv(output_file_path, sep=',', encoding='UTF-8', engine='python')
                df1.columns=headers
                df1.columns = df1.columns.str.replace('[', '').str.replace(']', '')
                if checkbox_var3:
                    df1 = df1[temp_sel1]
                else:
                    df1= df1[selection['File1']]
                df1 = df1.apply(lambda col: col.map(lambda x: x.replace('[', '').replace(']', '') if isinstance(x, str) else x))
                df1=remove_special_characters(df1)
                timecol=df1.columns[0]
                #print('start1')
                for col in df1.columns:
                    if col != 'Event' and col!= timecol:
                        for i in range(len(df1[col])):
                                
                            try:
                                # Try to convert the individual value to numeric, handling errors and replacing commas with dots
                                df1.at[i, col] = pd.to_numeric(str(df1.at[i, col]).replace(',', '.'), errors='coerce')

                                # Check if the value is non-numeric
                                if not pd.api.types.is_numeric_dtype(df1.at[i, col]):
                                    # If the value is non-numeric, replace it with NaN
                                    df1.at[i, col] = np.nan

                            except:
                                # Use a different approach to handle non-string values
                                df1.at[i,col] = float(df1.at[i, col]) if pd.api.types.is_numeric_dtype(df1.at[i, col]) else np.nan
                    else:
                        pass
                df1.to_csv('output0.csv', index=False)
            try: 

                df2.columns = df2.columns.str.replace('[', '').str.replace(']', '')
                
                df2= df2[selection['File2']]
                df2 = df2.apply(lambda col: col.map(lambda x: x.replace('[', '').replace(']', '') if isinstance(x, str) else x))
                timecol=df2.columns[0]
                for col in df2.columns:
                    if col != 'Event' and col!= timecol:
                        for i in range(len(df2[col])):
                            try:
                                # Try to convert the individual value to numeric, handling errors and replacing commas with dots
                                df2.at[i, col] = pd.to_numeric(str(df2.at[i, col]).replace(',', '.'), errors='coerce')

                                # Check if the value is non-numeric
                                if not pd.api.types.is_numeric_dtype(df2.at[i, col]):
                                    # If the value is non-numeric, replace it with NaN
                                    df2.at[i, col] = np.nan

                            except:
                                # Use a different approach to handle non-string values
                                df2.at[i,col] = float(df2.at[i, col]) if pd.api.types.is_numeric_dtype(df2.at[i, col]) else np.nan
                    else:
                        pass
                df2.to_csv('output1.csv', index=False)
            except Exception as e:           
                remove_spaces_and_replace_with_comma('output1.csv')
                file_name, file_extension = os.path.splitext('output1.csv')
                output_file_path = f"{file_name}_modified.csv"
                with open(output_file_path, 'r') as f:
                    reader = csv.reader(f, delimiter=',')
                    headers = next(reader)  # Read the first row (header)
                    # If the header row is empty, read the next non-empty row as the header
                    while not any(headers):
                        
                        headers = next(reader)
                for i, element in enumerate(headers):
                    headers[i] = element.replace('[', '').replace(']','')
                
                df2.to_csv('output1.csv', index=False)
                df2 = pd.read_csv(output_file_path, sep=',', encoding='UTF-8', engine='python')
                df2.columns=headers

                df2= df2[selection['File2']]
                
                df2 = df2.apply(lambda col: col.map(lambda x: x.replace('[', '').replace(']', '') if isinstance(x, str) else x))
                df2=remove_special_characters(df2)
                timecol=df2.columns[0]
                for col in df2.columns:
                    if col != 'Event' and col!= timecol:
                        for i in range(len(df2[col])):
                            # Try to convert the individual value to numeric, handling errors and replacing commas with dots        
                            try:
                                # Try to convert the individual value to numeric, handling errors and replacing commas with dots
                                df2.at[i, col] = pd.to_numeric(str(df2.at[i, col]).replace(',', '.'), errors='coerce')

                                # Check if the value is non-numeric
                                if not pd.api.types.is_numeric_dtype(df2.at[i, col]):
                                    # If the value is non-numeric, replace it with NaN
                                    df2.at[i, col] = np.nan

                            except:
                                # Use a different approach to handle non-string values
                                df2.at[i,col] = float(df2.at[i, col]) if pd.api.types.is_numeric_dtype(df2.at[i, col]) else np.nan
                    else:
                        pass
                df2.to_csv('output1.csv', index=False)
            check1,column_date1, format_date= is_date_column('output0.csv') #check1=True if date column is present in file1
            check2, column_date2,format_date1= is_date_column('output1.csv') #check2=True if date column is present in file2 

            df1 = pd.read_csv('output0.csv', sep=',', encoding='UTF-8', engine='python')
            df2 = pd.read_csv('output1.csv', sep=',', encoding='UTF-8', engine='python')

            count=0
            count1=0

            if check1:
                try:
                    time1 = pd.to_datetime(df1[column_date1], format=format_date)  # Convert the time column to datetime format
                    time1 = time1.strftime('%H:%M:%S') # Convert the time column to datetime format and then to string format
                    df1[column_date1]=time1 #assigning the time column to the dataframe
                                
                    for d in df1[column_date1]:

                        if str(d) == time1[0]:
                            count=count+ 1

                    if format_date=='%Y-%m-%d %H:%M' or format_date=='%d/%m/%Y %H:%M': 
                        #if the format is %Y-%m-%d %H:%M or %d/%m/%Y %H:%M 
                        #then the time is in hour and minutes format so we need to convert to add seconds

                        seconds_to_start = 60-count
                        seconds_to_add = 1
                        df1[column_date1] = insert_seconds_progressively(time1, seconds_to_start, seconds_to_add) #inserting seconds in the time column
                        df1[column_date1] = df1[column_date1].apply(lambda x: pd.to_datetime(x).strftime('%H:%M:%S'))

                        time1 = pd.to_datetime(df1[column_date1], format='%H:%M:%S')

                    format_date='%H:%M:%S'

                except AttributeError as e:
                    time1 = pd.to_datetime(time1)
                    time1 = time1.dt.strftime('%H:%M:%S')
                    df1[column_date1]=time1
                    
                    for d in df1[column_date1]:
                        if str(d) == time1[0]:
                            count=count+ 1
                    if format_date=='%Y-%m-%d %H:%M' or format_date=='%d/%m/%Y %H:%M':
                        seconds_to_start = 60-count
                        seconds_to_add = 1
                        df1[column_date1] = insert_seconds_progressively(time1, seconds_to_start, seconds_to_add)
                        df1[column_date1] = df1[column_date1].apply(lambda x: pd.to_datetime(x).strftime('%H:%M:%S'))

                        time1 = pd.to_datetime(df1[column_date1], format='%H:%M:%S')

                    format_date='%H:%M:%S'
            else:
                # Reference time 
                ref_time = pd.to_datetime(text_input4.get()) # Reference time input by the user
                column_date1=df1.columns[0]
                df1=apply_formulas_to_column(df1, ref_time, column_date1)

                format_date='%H:%M:%S'
                df1 = remove_special_characters(df1)
                df1.to_csv('output0.csv', index=False, encoding='UTF-8')

            if check2:
                try:
                    time2 = pd.to_datetime(df2[column_date2], format=format_date1) # Convert the time column to datetime format
                    time2 = time2.strftime('%H:%M:%S')

                    df2[column_date2]=time2
                    
                    for d in df2[column_date2]:
                        if str(d) == time2[0]:
                            count1=count1+ 1

                    if format_date1=='%Y-%m-%d %H:%M' or format_date1=='%d/%m/%Y %H:%M': 
                        #if the format is %Y-%m-%d %H:%M or %d/%m/%Y %H:%M then the time is in hour and minutes format so we need to convert to add seconds
                        seconds_to_start = 60-count1
                        seconds_to_add = 1
                        df2[column_date2] = insert_seconds_progressively(time2, seconds_to_start, seconds_to_add) #inserting seconds in the time column
                        df2[column_date2] = df2[column_date2].apply(lambda x: pd.to_datetime(x).strftime('%H:%M:%S'))

                        time2 = pd.to_datetime(df2[column_date2], format='%H:%M:%S')
                    
                    format_date1='%H:%M:%S'

                except AttributeError as e:
                    time2 = pd.to_datetime(time2)
                    time2 = time2.dt.strftime('%H:%M:%S')
                    df2[column_date2]=time2
                
                    for d in df2[column_date2]:
                        if str(d) == time2[0]:
                            count1=count1+ 1
                    if format_date1=='%Y-%m-%d %H:%M' or format_date1=='%d/%m/%Y %H:%M':
                
                        seconds_to_start = 60-count1
                        seconds_to_add = 1
                        df2[column_date2] = insert_seconds_progressively(time2, seconds_to_start, seconds_to_add)
                        df2[column_date2] = df2[column_date2].apply(lambda x: pd.to_datetime(x).strftime('%H:%M:%S'))

                        time2 = pd.to_datetime(df2[column_date2], format='%H:%M:%S')

                    format_date1='%H:%M:%S'
            else:

                # Reference time 
                ref_time = pd.to_datetime(text_input5.get()) # Reference time input by the user
            
                column_date2=df2.columns[0]
                df2=apply_formulas_to_column(df2, ref_time, column_date2)

                format_date1='%H:%M:%S'
                df2 = remove_special_characters(df2)
                
                df2.to_csv('output1.csv', index=False, encoding='UTF-8')

            for col in df1.columns: 
                if col==column_date1 or col=='Event':
                    pass
                else:
                    for i in range(len(df1[col])):        
                        try:
                            # Try to convert the individual value to numeric, handling errors and replacing commas with dots
                            df1.at[i, col] = pd.to_numeric(str(df1.at[i, col]).replace(',', '.'), errors='coerce')

                            # Check if the value is non-numeric
                            if not pd.api.types.is_numeric_dtype(df1.at[i, col]):
                                # If the value is non-numeric, replace it with NaN
                                df1.at[i, col] = np.nan

                        except:
                            # Use a different approach to handle non-string values
                            df1.at[i,col] = float(df1.at[i, col]) if pd.api.types.is_numeric_dtype(df1.at[i, col]) else np.nan

            for col in df2.columns:
                if col==column_date2 or col=='Event':
                    pass
                else:    
                    for i in range(len(df2[col])):
                        
                        try:
                            # Try to convert the individual value to numeric, handling errors and replacing commas with dots
                            df2.at[i, col] = pd.to_numeric(str(df2.at[i, col]).replace(',', '.'), errors='coerce')

                            # Check if the value is non-numeric
                            if not pd.api.types.is_numeric_dtype(df2.at[i, col]):
                                # If the value is non-numeric, replace it with NaN
                                df2.at[i, col] = np.nan

                        except:
                            # Use a different approach to handle non-string values
                            df2.at[i,col] = float(df2.at[i, col]) if pd.api.types.is_numeric_dtype(df2.at[i, col]) else np.nan


            df2.to_csv('output1.csv', index=False)
            df1.to_csv('output0.csv', index=False)

            if df1[column_date1].dtype == 'datetime64[ns]':
                time1 = df1[column_date1]
            else:
                time1 = datetime.strptime(df1[column_date1].iloc[0], format_date).time()

            if df2[column_date2].dtype == 'datetime64[ns]':
                time2 = df2[column_date2]
            else:
                time2 = datetime.strptime(df2[column_date2].iloc[0], format_date1).time()
            
            df1, df2, delta_seconds = align_dataframes(df1, df2, column_date1, column_date2)
            

            sec= delta_seconds
            if checkbox_var3: #if event = true
                
                indices_event=[]
                with open('options_event.txt') as f:
                    event_lines = f.read().splitlines()
                    for line in event_lines:
                        # Split each line at ':' and take the first part
                        parts = line.split('@#@')
                        if len(parts) > 0:
                            event_name = parts[1].strip()
                            # Check if the index is numeric
                            index = df1.index[df1['Event'] == event_name]
                            if not index.empty:
                                index = index.min()
                                index = int(index)
                                indices_event.append(index)

                    indices_event=sorted(indices_event)
                    print('indices')
                    print(indices_event)
        
                    if len(indices_event) == 1:
                        start_event=indices_event[0]
                        if sec:
                            stop_event = df1.index[-1-int(sec)]
                        else:
                            stop_event = df1.index[-1]

                    elif len(indices_event) == 2:
                        start_event=indices_event[0]
                        stop_event = indices_event[1]
                        
                    else:
                        start_event=df1.index[0]
                        stop_event = df1.index[-1]

                    if checkbox1 and select==2:
                        idx_start_line=int(start_event)#+sec
                        idx_stop_line=int(stop_event)#+sec

                    else: 
                        idx_start_line=int(start_event)
                        idx_stop_line=int(stop_event)

                    start_line=df1.loc[idx_start_line, df1.columns[0]]
                    stop_line = df1.loc[idx_stop_line, df1.columns[0]]

                    start_line_df1 = idx_start_line
                    stop_line_df1 = idx_stop_line+1
                    df1 = df1.iloc[start_line_df1:stop_line_df1]
            
                    try:
                        try:
                            start_line_df2 = df2.index[df2.iloc[:, 0] == start_line].tolist()[0]
                        except IndexError:
                            start_line_df2 = df2.index[0]
                        try:
                            stop_line_df2 = df2.index[df2.iloc[:, 0] == stop_line].tolist()[0]
                        except IndexError:
                            # Se non trova il valore, usa l'ultimo indice
                            stop_line_df2 = df2.index[-1]
                        stop_line_df2 = stop_line_df2 + 1

                        df2 = df2.iloc[start_line_df2:stop_line_df2]
                    except Exception:
                        df2 = df2.iloc[idx_start_line:idx_stop_line]

                    df1.to_csv('output0.csv', index=False, encoding='UTF-8')
                    df2.to_csv('output1.csv', index=False, encoding='UTF-8')
                    
            select_2columns(df1, df2)
                    

        elif checkbox1: 
            
            wb = load_workbook(output_file)
            ws = wb['Files']
            start_row = ws.max_row + 1
            for item in files_folder1:
                if item:
                    cell = ws.cell(row=start_row, column=1, value=item)
                start_row += 1
            wb.save(output_file)

            delimiter1 = auto_detect_delimiter('output0.csv')
            df1 = pd.read_csv('output0.csv', sep=',', encoding='UTF-8', engine='python')       
            df1.columns = df1.columns.str.replace('[', '').str.replace(']', '')
            df1=remove_special_characters(df1)

            try: 
                df1 = df1.applymap(lambda x: x.replace('[', '').replace(']', '') if isinstance(x, str) else x)

                if checkbox_var3:
                    df1 = df1[temp_sel1]
                else:
                    df1= df1[selection['File1']]
                
                df1.to_csv('output0.csv', index=False)
                df1 = df1.loc[:, ~df1.columns.duplicated()]
                first_column = True  # Set a flag to identify the first column
                for col in df1.columns:
                    if first_column or col == 'Event':
                        first_column = False  # Set the flag to False after processing the first column
                        continue  # Skip the first column and the 'Event' column

                    for i in range(len(df1[col])):
                        try:
                            # Try to convert the individual value to numeric, handling errors and replacing commas with dots
                            df1.at[i, col] = pd.to_numeric(str(df1.at[i, col]).replace(',', '.'), errors='coerce')

                            # Check if the value is non-numeric
                            if not pd.api.types.is_numeric_dtype(df1.at[i, col]):
                                # If the value is non-numeric, replace it with NaN
                                df1.at[i, col] = np.nan

                        except:
                            # Use a different approach to handle non-string values
                            df1.at[i,col] = float(df1.at[i, col]) if pd.api.types.is_numeric_dtype(df1.at[i, col]) else np.nan
                df1.to_csv('output0.csv', index=False)
                
            except Exception as e:
                
                df1 = pd.read_csv('output0.csv', sep=',', encoding='UTF-8', engine='python')
                df1=remove_special_characters(df1)
                df1.to_csv('output0.csv', index=False)
                remove_spaces_and_replace_with_comma('output0.csv')
                file_name, file_extension = os.path.splitext('output0.csv')
                output_file_path = f"{file_name}_modified.csv"
                with open(output_file_path, 'r') as f:
                    reader = csv.reader(f, delimiter=',')
                    headers = next(reader)  # Read the first row (header)

                    # If the header row is empty, read the next non-empty row as the header
                    while not any(headers):
                        
                        headers = next(reader)

                headers = [col for col in headers if col != '']

                df1 = pd.read_csv(output_file_path, sep=',', encoding='UTF-8', engine='python', usecols=headers)

                df1.columns = [col.replace('[', '').replace(']', '') for col in df1.columns]
                df1 = df1.applymap(lambda x: x.replace('[', '').replace(']', '') if isinstance(x, str) else x)
                df1=remove_special_characters(df1)
                print(df1.columns)
                if checkbox_var3:
                    df1= df1[temp_sel1]
                else:
                    df1= df1[selection['File1']]
                df1=remove_special_characters(df1)
                first_column = True  # Set a flag to identify the first column

                for col in df1.columns:
                    if first_column or col == 'Event':
                        first_column = False  # Set the flag to False after processing the first column
                        continue  # Skip the first column and the 'Event' column

                    for i in range(len(df1[col])):
                        try:
                            # Try to convert the individual value to numeric, handling errors and replacing commas with dots
                            df1.at[i, col] = pd.to_numeric(str(df1.at[i, col]).replace(',', '.'), errors='coerce')

                            # Check if the value is non-numeric
                            if not pd.api.types.is_numeric_dtype(df1.at[i, col]):
                                # If the value is non-numeric, replace it with NaN
                                df1.at[i, col] = np.nan

                        except:
                            # Use a different approach to handle non-string values
                            df1.at[i,col] = float(df1.at[i, col]) if pd.api.types.is_numeric_dtype(df1.at[i, col]) else np.nan
                df1.to_csv('output0.csv', index=False)
            check1,column_date1, format_date= is_date_column('output0.csv')
            df1 = pd.read_csv('output0.csv', sep=',', encoding='UTF-8', engine='python')
            
            count=0
            
            if check1:
                try:
                    time1 = pd.to_datetime(df1[column_date1], format=format_date)
                    
                    time1 = time1.strftime('%H:%M:%S')
                    
                    df1[column_date1]=time1
                    
                    for d in df1[column_date1]:
                    
                        if str(d) == time1[0]:
                            count=count+ 1
                    
                    if format_date=='%Y-%m-%d %H:%M' or format_date=='%d/%m/%Y %H:%M':
                        
                        seconds_to_start = 60-count
                        seconds_to_add = 1
                        df1[column_date1] = insert_seconds_progressively(time1, seconds_to_start, seconds_to_add)
                        df1[column_date1] = df1[column_date1].apply(lambda x: pd.to_datetime(x).strftime('%H:%M:%S'))

                        time1 = pd.to_datetime(df1[column_date1], format='%H:%M:%S')
                        
                    format_date='%H:%M:%S'
                except AttributeError as e:
                    time1 = pd.to_datetime(time1)
                    time1 = time1.dt.strftime('%H:%M:%S')
                
                    df1[column_date1]=time1
                    
                
                    for d in df1[column_date1]:
                    
                        if str(d) == time1[0]:
                            count=count+ 1
                
                    if format_date=='%Y-%m-%d %H:%M' or format_date=='%d/%m/%Y %H:%M':
                    
                        seconds_to_start = 60-count
                        seconds_to_add = 1
                        df1[column_date1] = insert_seconds_progressively(time1, seconds_to_start, seconds_to_add)
                        df1[column_date1] = df1[column_date1].apply(lambda x: pd.to_datetime(x).strftime('%H:%M:%S'))

                        time1 = pd.to_datetime(df1[column_date1], format='%H:%M:%S')
                    
                    format_date='%H:%M:%S'
            else:

                # Reference time 
                ref_time = pd.to_datetime(text_input4.get()) 
                column_date1=df1.columns[0]
                df1=apply_formulas_to_column(df1, ref_time, column_date1)

                format_date='%H:%M:%S'
                df1 = remove_special_characters(df1)
                print(df1)
                print('prob qui')
                df1.to_csv('output0.csv', index=False, encoding='UTF-8')

            for col in df1.columns:
                if col==column_date1 or col=='Event':
                    pass
                else:
                    for i in range(len(df1[col])):
                        try:
                            # Try to convert the individual value to numeric, handling errors and replacing commas with dots
                            df1.at[i, col] = pd.to_numeric(str(df1.at[i, col]).replace(',', '.'), errors='coerce')

                            # Check if the value is non-numeric
                            if not pd.api.types.is_numeric_dtype(df1.at[i, col]):
                                #print(f"Non-numeric value found at index {i}.")
                                # If the value is non-numeric, replace it with NaN
                                df1.at[i, col] = np.nan

                        except:
                            # Use a different approach to handle non-string values
                            df1.at[i,col] = float(df1.at[i, col]) if pd.api.types.is_numeric_dtype(df1.at[i, col]) else np.nan
            
            df1.to_csv('output0.csv', index=False)
            if df1[column_date1].dtype == 'datetime64[ns]':
                time1 = df1[column_date1]
            else:
                time1 = datetime.strptime(df1[column_date1].iloc[0], format_date).time()

            if checkbox_var3:
                
                indices_event=[]
                with open('options_event.txt') as f:
                    event_lines = f.read().splitlines()
                    for line in event_lines:
                        print(line)
                        # Split each line at ':' and take the first part
                        parts = line.split('@#@')
                        print(parts)
                        if len(parts) > 0:
                            event_name = parts[1].strip()
                            # Check if the index is numeric
                            index = df1.index[df1['Event'] == event_name]                                            
                            if not index.empty:
                                index = index.min()
                                index = int(index)
                                indices_event.append(index)

                    indices_event=sorted(indices_event)   
                    if len(indices_event) == 1:
                        start_event=indices_event[0]
                        stop_event = df1.index[-1]
                    elif len(indices_event) == 2:
                        start_event=indices_event[0]
                        stop_event = indices_event[1]
                    else:
                        start_event=df1.index[0]
                        stop_event = df1.index[-1]
                    
                    if checkbox1 and select==2:
                        idx_start_line=int(start_event)+sec
                        idx_stop_line=int(stop_event)+sec
                    else: 
                        idx_start_line=int(start_event)
                        idx_stop_line=int(stop_event)

                    start_line=df1.loc[idx_start_line, df1.columns[0]]
                    stop_line=df1.loc[idx_stop_line, df1.columns[0]]
                    start_line_df1 = idx_start_line
                    stop_line_df1 = idx_stop_line+1
                    df1 = df1.iloc[start_line_df1:stop_line_df1]
                    df1.to_csv('output0.csv', index=False, encoding='UTF-8')

            select_columns(df1)

        if checkbox:
            checkp=0
            if not checkbox1:
                wb = load_workbook(output_file)
                ws = wb['Files']
                start_row = ws.max_row + 1
                for item in files_folder1:
                    if item:
                        cell = ws.cell(row=start_row, column=1, value=item)
                    
                    start_row += 1
                wb.save(output_file)
                
                df1 = pd.read_csv('output0.csv', sep=',', encoding='UTF-8', engine='python')
                delimiter1 = auto_detect_delimiter('output0.csv')
            
                df1 = pd.read_csv('output0.csv', sep=delimiter1, encoding='UTF-8', engine='python')
                
                df1.columns = df1.columns.str.replace('[', '').str.replace(']', '')

                df1=remove_special_characters(df1)
            
                try: 
                    df1 = df1.applymap(lambda x: x.replace('[', '').replace(']', '') if isinstance(x, str) else x)
                    df1= df1[selection['File1']]
                                
                    first_column = True  # Set a flag to identify the first column

                    for col in df1.columns:
                        if len(df1.columns)>1:
                            if first_column or col == 'Event':
                                first_column = False  # Set the flag to False after processing the first column
                                continue  # Skip the first column and the 'Event' column

                        for i in range(len(df1[col])):
                            try:
                                # Try to convert the individual value to numeric, handling errors and replacing commas with dots
                                df1.at[i, col] = pd.to_numeric(str(df1.at[i, col]).replace(',', '.'), errors='coerce')

                                # Check if the value is non-numeric
                                if not pd.api.types.is_numeric_dtype(df1.at[i, col]):
                                    # If the value is non-numeric, replace it with NaN
                                    df1.at[i, col] = np.nan

                            except:
                                # Use a different approach to handle non-string values
                                df1.at[i,col] = float(df1.at[i, col]) if pd.api.types.is_numeric_dtype(df1.at[i, col]) else np.nan
                    df1.to_csv('output0.csv', index=False)
                    
                except Exception as e:
                    
                    remove_spaces_and_replace_with_comma('output0.csv')
                    file_name, file_extension = os.path.splitext('output0.csv')
                    output_file_path = f"{file_name}_modified.csv"
                    with open(output_file_path, 'r') as f:
                        reader = csv.reader(f, delimiter=',')
                        headers = next(reader)  # Read the first row (header)

                        # If the header row is empty, read the next non-empty row as the header
                        while not any(headers):
                            
                            headers = next(reader)
                    for i, element in enumerate(headers):
                        headers[i] = element.replace('[', '').replace(']','')
                    
                    df1.to_csv('output0.csv', index=False)
                
                    df1 = pd.read_csv(output_file_path, sep=',', encoding='UTF-8', engine='python')

                    df1.columns=headers
                    
                    df1= df1[selection['File1']]
                    df1 = df1.applymap(lambda x: x.replace('[', '').replace(']', '') if isinstance(x, str) else x)
                    df1=remove_special_characters(df1)
                    first_column = True  # Set a flag to identify the first column

                    for col in df1.columns:
                        if len(df1.columns)>1:
                            if first_column or col == 'Event':
                                first_column = False  # Set the flag to False after processing the first column
                                continue  # Skip the first column and the 'Event' column
                        for i in range(len(df1[col])):
                            try:
                                # Try to convert the individual value to numeric, handling errors and replacing commas with dots
                                df1.at[i, col] = pd.to_numeric(str(df1.at[i, col]).replace(',', '.'), errors='coerce')

                                # Check if the value is non-numeric
                                if not pd.api.types.is_numeric_dtype(df1.at[i, col]):
                                    #print(f"Non-numeric value found at index {i}.")
                                    # If the value is non-numeric, replace it with NaN
                                    df1.at[i, col] = np.nan

                            except:
                                # Use a different approach to handle non-string values
                                df1.at[i,col] = float(df1.at[i, col]) if pd.api.types.is_numeric_dtype(df1.at[i, col]) else np.nan

                    df1.to_csv('output0.csv', index=False)
                check1,column_date1, format_date= is_date_column('output0.csv')
                df1 = pd.read_csv('output0.csv', sep=',', encoding='UTF-8', engine='python')           
                count=0

                if check1:
                    
                    try:
                        time1 = pd.to_datetime(df1[column_date1], format=format_date)
                        
                        time1 = time1.strftime('%H:%M:%S')
                        
                        df1[column_date1]=time1
                        

                        for d in df1[column_date1]:
                        
                            if str(d) == time1[0]:
                                count=count+ 1
                        
                        if format_date=='%Y-%m-%d %H:%M' or format_date=='%d/%m/%Y %H:%M':
                            
                            seconds_to_start = 60-count
                            seconds_to_add = 1
                            df1[column_date1] = insert_seconds_progressively(time1, seconds_to_start, seconds_to_add)
                            df1[column_date1] = df1[column_date1].apply(lambda x: pd.to_datetime(x).strftime('%H:%M:%S'))

                            time1 = pd.to_datetime(df1[column_date1], format='%H:%M:%S')
                        
                        format_date='%H:%M:%S'
                    except AttributeError as e:
                        time1 = pd.to_datetime(time1)
                        time1 = time1.dt.strftime('%H:%M:%S')
                    
                        df1[column_date1]=time1
                    
                        for d in df1[column_date1]:
                        
                            if str(d) == time1[0]:
                                count=count+ 1
                        
                        if format_date=='%Y-%m-%d %H:%M' or format_date=='%d/%m/%Y %H:%M':
                        
                            seconds_to_start = 60-count
                            seconds_to_add = 1
                            df1[column_date1] = insert_seconds_progressively(time1, seconds_to_start, seconds_to_add)
                            df1[column_date1] = df1[column_date1].apply(lambda x: pd.to_datetime(x).strftime('%H:%M:%S'))

                            time1 = pd.to_datetime(df1[column_date1], format='%H:%M:%S')
                        
                        format_date='%H:%M:%S'
            
                df1.to_csv('output0.csv', index=False)
            
            if checkbox_var3 and not checkbox1:

                indices_event=[]
                with open('options_event.txt') as f:
                    event_lines = f.read().splitlines()
                    for line in event_lines:
                        # Split each line at ':' and take the first part
                        parts = line.split('@#@')
                        if len(parts) > 0:
                            event_name = parts[1].strip()
                            # Check if the index is numeric
                            index = df1.index[df1['Event'] == event_name]                      

                            if not index.empty:
                                index = index.min()
                                index = int(index)
                                indices_event.append(index)

                    indices_event=sorted(indices_event) 
                    if len(indices_event) == 1:
                        start_event=indices_event[0]
                        stop_event = df1.index[-1]
                    elif len(indices_event) == 2:
                        start_event=indices_event[0]
                        stop_event = indices_event[1]
                    else:
                        start_event=df1.index[0]
                        stop_event = df1.index[-1]
                    
                    if checkbox1 and select==2:
                        idx_start_line=int(start_event)#+sec
                        idx_stop_line=int(stop_event)#+sec
                    else: 
                        idx_start_line=int(start_event)
                        idx_stop_line=int(stop_event)

                    start_line=df1.loc[idx_start_line, df1.columns[0]]
                    stop_line=df1.loc[idx_stop_line, df1.columns[0]]
            
                    start_line_df1 = idx_start_line
                    stop_line_df1 = idx_stop_line
                    df1 = df1.iloc[start_line_df1:stop_line_df1]
                    df1.to_csv('output0.csv', index=False, encoding='UTF-8')
                
            delimiter = auto_detect_delimiter('output0.csv')
            for i, element in enumerate(selection['File1']):
                    selection['File1'][i] = element.replace('[', '').replace(']','')
            for i, element in enumerate(selection['File2']):
                    selection['File2'][i] = element.replace('[', '').replace(']','')
            df = pd.read_csv('output0.csv', sep=',', encoding='UTF-8', engine='python')
            
            try:
                with open('output0.csv', 'r') as f:
                    reader = csv.reader(f, delimiter=',')
                    headers = next(reader)  # Read the first row (header)

                    # If the header row is empty, read the next non-empty row as the header
                    while not any(headers):
                        
                        headers = next(reader)
                    for i, element in enumerate(headers):
                        headers[i] = element.replace('[', '').replace(']','')
                
            except Exception as e:
                
                remove_spaces_and_replace_with_comma('output0.csv')
                file_name, file_extension = os.path.splitext('output0.csv')
                output_file_path = f"{file_name}_modified.csv"
                with open(output_file_path, 'r') as f:
                    reader = csv.reader(f, delimiter=',')
                    headers = next(reader)  # Read the first row (header)

                    # If the header row is empty, read the next non-empty row as the header
                    while not any(headers):
                        
                        headers = next(reader)
                for i, element in enumerate(headers):
                    headers[i] = element.replace('[', '').replace(']','')
                df = pd.read_csv(output_file_path, sep=',', encoding='UTF-8', engine='python')

            if checkbox1:
                selection['File1']=headers

            df.columns=headers
            df = remove_special_characters(df)
            col_header=selection['File1']

            for i, element in enumerate(col_header):
                        col_header[i] = element.replace('[', '').replace(']','')
            
            col_header = [item for item in col_header if item != 'Event']
        
            try: 
                temp=df[col_header]  # Seleziona solo le colonne desiderate
                
            except KeyError:
                df.columns = df.iloc[0].tolist()
                df.columns = df.columns.astype(str)
                temp=df[col_header]
            temp.to_csv('check.csv', index=False)

            if select==2:
                try:
                    with open('output1.csv', 'r') as f:
                        reader = csv.reader(f, delimiter=delimiter)
                        headers = next(reader)  # Read the first row (header)

                        # If the header row is empty, read the next non-empty row as the header
                        while not any(headers):
                            
                            headers = next(reader)
                        for i, element in enumerate(headers):
                            headers[i] = element.replace('[', '').replace(']','')
                    df0 = pd.read_csv('output1.csv', sep=',', encoding='UTF-8', engine='python')
                except Exception as e:
                    remove_spaces_and_replace_with_comma('output1.csv')
                    file_name, file_extension = os.path.splitext('output1.csv')
                    output_file_path = f"{file_name}_modified.csv"
                    with open(output_file_path, 'r') as f:
                        reader = csv.reader(f, delimiter=',')
                        headers = next(reader)  # Read the first row (header)

                        # If the header row is empty, read the next non-empty row as the header
                        while not any(headers):
                            
                            headers = next(reader)
                    for i, element in enumerate(headers):
                        headers[i] = element.replace('[', '').replace(']','')
                    
                    df0 = pd.read_csv(output_file_path, sep=',', encoding='UTF-8', engine='python')
                df0.columns=headers
                df0 = pd.read_csv('output1.csv')
                df0 = remove_special_characters(df0)
                
                selection['File2']=[]
                selection['File2']=df0.columns
            
                df2=df0[selection['File2']]
                df=pd.concat([df, df2], axis=1)
                col=selection['File2']
                df.to_csv('check.csv', index=False)
                col_header.extend(col)
            
            check1, format_date, colu=is_date_column2('check.csv')
            
            if check1:
                colu = list(colu)
                temp=df[colu]
                df = df.drop(columns=colu)
                for col_name in colu:
                    if col_name in col_header:    
                        try:              
                            col_header = col_header.drop(col_name, axis=1)
                        except AttributeError:
                            col_header.remove(col_name)
            if check1:
                df=pd.concat([df, temp], axis=1)
            
            if not checkbox1:
                select_columns(df)
            
            sheet_name = 'Data'+str(currentpage)
            df = pd.read_excel(output_file, sheet_name=sheet_name)

            wb = openpyxl.load_workbook(output_file)
            ws = wb[sheet_name]
            if minimum:
                lower_limit = float(minimum)
            else:
                lower_limit = 0.0
                minimum=0.0
            if maximum:
                upper_limit = float(maximum)
            else:
                upper_limit= 0.0
                maximum=0.0
            if not threshold:
                threshold=1
            out_of_range_rows = {col:[] for col in range(df.shape[1])}
                    
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid') 

            green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            
            row_index = currentpage + 2  # Example row index (1-based)
            column_index = 3
            fake_event=0
            if 'Event' in df.columns:
                df0_keepevent=df['Event']
                df = df.drop('Event', axis=1)
                fake_event=1
                        
            if var_unit2.get()=="Fail":
                header_high=[]
                for index, row in df.iterrows():
                    for col in range(len(row)):
                        #try:
                        if df.columns[col] == "Event":
                            continue
                        if isinstance(row[col], str) and ',' in row[col]:
                            value = float(row[col].replace(',', '.'))
                        if isinstance(row[col], str) and '.' in row[col]:
                            value = float(row[col])
                        if isinstance(row[col], str) and not ',' in row[col]:
                            value = np.nan
                        else:
                            value = float(row[col])
                        if var_unit3.get()=="x<Limit1 or x>Limit2":
                            if value < lower_limit or value > upper_limit:
                                out_of_range_rows[col].append(index+2)
                        elif var_unit3.get()=="Limit1<x<Limit2":
                            if lower_limit < value < upper_limit:
                                out_of_range_rows[col].append(index+2)
                        elif var_unit3.get()=="x<Limit1":
                            if value < lower_limit:
                                out_of_range_rows[col].append(index+2)
                        elif var_unit3.get()=="x>Limit1":
                            if value > lower_limit:
                                out_of_range_rows[col].append(index+2)
                        elif var_unit3.get() == "x<=Limit1":
                            if value <= lower_limit:
                                out_of_range_rows[col].append(index+2)
                        elif var_unit3.get() == "x>=Limit1":
                            if value >= lower_limit:
                                out_of_range_rows[col].append(index+2)
                        elif var_unit3.get() == "Limit1<=x<=Limit2":
                            if lower_limit <= value <= upper_limit:
                                out_of_range_rows[col].append(index+2)
                        elif var_unit3.get() == "x=Limit1":
                            if value == lower_limit:
                                out_of_range_rows[col].append(index+2)
                        
                # Group highlighting logic
                highlight_groups = {col:[] for col in out_of_range_rows.keys()}

                for col, indices in out_of_range_rows.items():
                    for i in indices:
                        if not highlight_groups[col]:
                            highlight_groups[col].append([i]) 
                        else:
                            last_idx = highlight_groups[col][-1][-1]
                            if i > last_idx+1:
                                highlight_groups[col].append([i])
                            else:
                                highlight_groups[col][-1].append(i)
                for col, groups in highlight_groups.items():
                        for group in groups:
                            if len(group) >= int(threshold):
                                for row_idx in group:
                                    ws.cell(row=row_idx, column=col+1+ fake_event).fill = red_fill
                                    header_high.append(df.columns[col]) 
                new_value = ""
                unique_list = []
                for item in header_high:
                    if item not in unique_list:
                        unique_list.append(item)
                if unique_list:
                    new_value = "Fail on " + ", ".join(str(x) for x in unique_list)               
                else:
                    new_value ='Pass'
            
            if var_unit2.get()=="Pass":
                header_high=[]
                for index, row in df.iterrows():
                    for col in range(len(row)):
                        #try:
                        if isinstance(row[col], str) and ',' in row[col]:
                            value = float(row[col].replace(',', '.'))
                        if isinstance(row[col], str) and '.' in row[col]:
                            value = pd.to_numeric(row[col])
                        if isinstance(row[col], str) and not ',' in row[col]:
                            value = np.nan
                        else:
                            value = float(row[col])
                        if var_unit3.get()=="x<Limit1 or x>Limit2":
                            if value < lower_limit or value > upper_limit:
                                out_of_range_rows[col].append(index+2)
                        elif var_unit3.get()=="Limit1<x<Limit2":
                            if lower_limit < value < upper_limit:
                                out_of_range_rows[col].append(index+2)
                        elif var_unit3.get()=="x<Limit1":
                            if value < lower_limit:
                                out_of_range_rows[col].append(index+2)
                        elif var_unit3.get()=="x>Limit1":
                            if value > lower_limit:
                                out_of_range_rows[col].append(index+2)
                        elif var_unit3.get() == "x<=Limit1":
                            if value <= lower_limit:
                                out_of_range_rows[col].append(index+2)
                        elif var_unit3.get() == "x>=Limit1":
                            if value >= lower_limit:
                                out_of_range_rows[col].append(index+2)
                        elif var_unit3.get() == "Limit1<=x<=Limit2":
                            if lower_limit <= value <= upper_limit:
                                out_of_range_rows[col].append(index+2)
                        elif var_unit3.get() == "x=Limit1":
                            if value == lower_limit:
                                out_of_range_rows[col].append(index+2)

                # Group highlighting logic
                highlight_groups = {col:[] for col in out_of_range_rows.keys()}

                for col, indices in out_of_range_rows.items():
                    for i in indices:
                        if not highlight_groups[col]:
                            highlight_groups[col].append([i]) 
                        else:
                            last_idx = highlight_groups[col][-1][-1]
                            if i > last_idx+1:
                                highlight_groups[col].append([i])
                            else:
                                highlight_groups[col][-1].append(i)
                new_value = ""
                header_high=[]
                for col, groups in highlight_groups.items():
                    for group in groups:
                        if len(group) >= int(threshold):
                            for row_idx in group:
                                ws.cell(row=row_idx, column=col+1+fake_event).fill = green_fill
                                header_high.append(df.columns[col]) 

                unique_list = []
                for item in header_high:
                    if item not in unique_list:
                        unique_list.append(item)
                if unique_list:
                    new_value = "Pass on " + ", ".join(str(x) for x in unique_list)              
                else:
                    new_value ='Fail'
            
            sheet1 = wb['Files']
            sheet1.cell(row=row_index, column=column_index, value=new_value)

            wb.save(output_file)

        if checkbox2:
            checkp=0
            
            df0 = pd.read_csv('backup.csv', sep=',', engine='python')
            
            df0 = remove_special_characters(df0)
            temp_sel1=remove_special_characters_from_list(temp_sel1)
            temp_sel2=remove_special_characters_from_list(temp_sel2)

            df= df0[temp_sel1]
        
            if select==2:
            
                df2=df0[temp_sel2]
                df=pd.concat([df, df2], axis=1)
                
            df = df.loc[:, ~df.columns.duplicated()]

            lung = max(len(df.iloc[:, i]) for i in range(len(df.columns)))

            check1, format_date, colu=is_date_column2('backup.csv')

            if colu:
                unique_colu, index = np.unique(colu, return_index=True)
                colu = pd.Series(unique_colu)
                colu.drop_duplicates(inplace=True)
                # Sorting based on the original indices
                colu = colu.sort_index(ascending=True)  # or ascending=False for descending order
            
            if check1 and len(colu)==2:
                print("Performing date check to plot")
                print(selection['File1'])
                print(selection['File2'])
                print(df.columns)
                try:
                    df_plot1=df.drop(columns=selection['File1']) #cambio da df0.drop a df.drop
                except:
                    if "Event.1" in selection['File1']:
                        selection['File1'].remove("Event.1")
                    df_plot1=df.drop(columns=selection['File1'])
                
                df_plot2=df.drop(columns=selection['File2'])
                
                df0 = df_plot1
                df2 = df_plot2     

                colx1 = df0.iloc[:,0]
                colx2 = df2.iloc[:,0]

                
                df_colx1 = pd.DataFrame({'colx': colx1})
                df_colx2 = pd.DataFrame({'colx': colx2})

                df0 = df0.drop(columns=df0.columns[0])
                df2 = df2.drop(columns=df2.columns[0])
                
                for col in df0.columns:
                    if df0[col].dtype == 'object':
                        df0[col] = df0[col].str.replace(',', '.')
                for col in df2.columns:
                    if df2[col].dtype == 'object':
                        df2[col] = df2[col].str.replace(',', '.')
                
                if checkbox:

                    if var_unit3.get()=="x < Limit1 or x > Limit2" or var_unit3.get()=="Limit1<x<Limit2":
                        maxp=pd.Series([float(maximum)]*lung, name= 'Limit2')
                        maxp.index = df0.index
                        df0=pd.concat([df0, maxp], axis=1)

                    minp1=pd.Series([float(minimum)]*lung, name= 'Limit1')                
                    minp1.index = df0.index

                    df0 = df0.apply(pd.to_numeric, errors='coerce')               

                    if var_unit3.get()=="x < Limit1 or x > Limit2" or var_unit3.get()=="Limit1<x<Limit2":
                        maxp=pd.Series([float(maximum)]*lung, name= 'Limit2')
                        maxp.index = df2.index
                        df2=pd.concat([df2, maxp], axis=1)

                    minp2=pd.Series([float(minimum)]*lung, name= 'Limit1')               
                    minp2.index = df2.index

                    df2 = df2.apply(pd.to_numeric, errors='coerce') 

                    if len(minp1)>len(minp2):
                        df0=pd.concat([df0, minp1], axis=1)
                    else:
                        df2=pd.concat([df2, minp2], axis=1)
                    
                df0=pd.concat([df_colx1, df0], axis=1)
                df2=pd.concat([df_colx2, df2], axis=1)
                
                df0['colx'] = df0['colx'].str.extract(r'(\d{2}:\d{2}:\d{2})')
                df0['colx'] = pd.to_datetime(df0['colx'], format='%H:%M:%S').dt.time

                df0.to_csv('prova2.csv', index=False)

                columns_to_plot1 = df0.columns                    
                check1, format_date, colu=is_date_column2('prova2.csv')

                columns_to_plot1 = [col for col in columns_to_plot1 if col != 'colx']
                columns_to_plot1 = [col for col in columns_to_plot1 if col != 'Event']
                columns_to_plot1 = [col for col in columns_to_plot1 if col not in colu]

                labels1 = [str(column) for column in columns_to_plot1]
                # Set up the plot
                df0['colx'] = df0['colx'].astype(str)
                df0['colx'] = df0['colx'].str.split('.').str[0]

                df0 = df0.loc[:, ~df0.columns.duplicated()]   

                df2['colx'] = df2['colx'].str.extract(r'(\d{2}:\d{2}:\d{2})')
                df2['colx'] = pd.to_datetime(df2['colx'], format='%H:%M:%S').dt.time
                df2.to_csv('prova2.csv', index=False)

                columns_to_plot2 = df2.columns                    
                check1, format_date, colu=is_date_column2('prova2.csv')

                columns_to_plot2 = [col for col in columns_to_plot2 if col != 'colx']
                columns_to_plot2 = [col for col in columns_to_plot2 if col != 'Event']
                columns_to_plot2 = [col for col in columns_to_plot2 if col not in colu]

                labels2 = [str(column) for column in columns_to_plot2]
                # Set up the plot
                df2['colx'] = df2['colx'].astype(str)
                df2['colx'] = df2['colx'].str.split('.').str[0]
                df2 = df2.loc[:, ~df2.columns.duplicated()]     

                #####       
                # Convert 'colx' to datetime format
                df0['colx'] = pd.to_datetime(df0['colx'])
                df2['colx'] = pd.to_datetime(df2['colx'])

                # Identify the first non-00:00:00 time
                df0['colx'] = df0['colx'].replace('', pd.NaT)
                df0['colx'] = pd.to_datetime(df0['colx'], format='%H:%M:%S', errors='coerce')

                fig, ax = plt.subplots() 
                if checkbox:
                    if 'Limit1' in columns_to_plot1:
                        columns_to_plot1.remove('Limit1') 
                    if 'Limit2' in columns_to_plot1:
                        columns_to_plot1.remove('Limit2')

                # Plot data for df0
                for column in columns_to_plot1:
                    df0 = df0.dropna(subset=column)
                    ax.plot(df0['colx'], df0[column], label=str(column))

                if checkbox:
                    if 'Limit1' in columns_to_plot2:
                        columns_to_plot2.remove('Limit1') 
                    if 'Limit2' in columns_to_plot2:
                        columns_to_plot2.remove('Limit2')
                
                df2['colx'] = df2['colx'].replace('', pd.NaT)
                df2['colx'] = pd.to_datetime(df2['colx'], format='%H:%M:%S', errors='coerce')

                for column in columns_to_plot2:
                    df2 = df2.dropna(subset=column)
                    ax.plot(df2['colx'], df2[column], label=str(column))
                
                if minimum:
                    ax.axhline(y=float(minimum), linestyle='--', label='Reference Limit1', color='black')
                if maximum:
                    if var_unit3.get()=="x<Limit1 or x>Limit2" or var_unit3.get()=="Limit1<x<Limit2" or var_unit3.get()=="Limit1<=x<=Limit2":
                        ax.axhline(y=float(maximum), linestyle='--', label='Reference Limit2', color='black')

                ax.xaxis.set_major_locator(plt.MaxNLocator(6))
                ax.yaxis.set_major_locator(plt.MaxNLocator(6))

                # Format x-axis ticks to display only the hour, minute, and second
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))

                title = f'Values: {", ".join(str(x) for x in columns_to_plot1)}, {", ".join(str(x) for x in columns_to_plot2)}'
                plt.title(title)

                ax.legend(loc='upper left', bbox_to_anchor=(1.05, 1))
                ax.set_xlabel('Samples')
                ax.set_ylabel('Values in ' + var_unit.get())

                columns_to_plot= columns_to_plot1+columns_to_plot2
                # Remove duplicates from columns_to_plot
                columns_to_plot = list(set(columns_to_plot))

                # Remove 'min' and 'max' if present and checkbox is checked
                if checkbox:
                    for col in ['Limit1', 'Limit2']:
                        if col in columns_to_plot:
                            columns_to_plot.remove(col)

                # Save the figure and close
                fig.savefig('plot.png', bbox_inches='tight')
                plt.close(fig)
                                
                img = Image.open('plot.png')
                excel_img = ExcelImage(img)
                # Load the workbook using openpyxl
                book = openpyxl.load_workbook(output_file)
                
                sheet_name = 'Data'+str(currentpage)  # Replace with the actual sheet name
                sheet = book[sheet_name]
                # Trova prima colonna vuota
                empty_col=len(selection['File1']) + len(selection['File2']) + 2

                # Aggiungi immagine nella colonna vuota, riga 1 
                sheet.add_image(excel_img, f'{openpyxl.utils.get_column_letter(empty_col)}1')  
                book.save(output_file)
                checkp=1
                # print("launching interactive plot")
                # launch_interactive_plot(df0,df2)

            ##arrivato qui
            elif check1:
                for col in colu.values:
                        
                    colx = df[col]#.squeeze().str.split(' ').str[1]
                    
                    df0 = pd.DataFrame({'colx': colx}) #cambiato da df0 a df
                    print(df0.columns)
                    df = df.drop(columns=colu) #cambiato da df0 a df
                    print(df.columns)
                    #print(df['Event'])
                    if 'Event' in df.columns:
                        df0_keepevent=df['Event']
                    df = df.apply(pd.to_numeric, errors='coerce')

                    for col in df.columns:
                        if df[col].dtype == 'object':
                            df[col] = df[col].str.replace(',', '.')
                    if checkbox:
                        if var_unit3.get()=="x < Limit1 or x > Limit2" or var_unit3.get()=="Limit1<x<Limit2":
                            maxp=pd.Series([float(maximum)]*lung, name= 'Limit2')
                            maxp.index = df.index
                            df=pd.concat([df, maxp], axis=1)
                        minp=pd.Series([float(minimum)]*lung, name= 'Limit1')
                        
                        minp.index = df.index
                        df=pd.concat([df, minp], axis=1)
                        
                        df = df.apply(pd.to_numeric, errors='coerce')
                        df.to_csv('prova1.csv', index=False)
                                
                    df=pd.concat([df0, df], axis=1) #commentato questo perchè tolto df0 riga 2015
                    df['colx'] = df['colx'].str.extract(r'(\d{2}:\d{2}:\d{2})')
                    df.to_csv('prova2.csv', index=False)
                    df['colx'] = pd.to_datetime(df['colx'], format='%H:%M:%S').dt.time
                    
                    columns_to_plot = df.columns               
                    check1, format_date, colu=is_date_column2('prova2.csv')

                    columns_to_plot = [col for col in columns_to_plot if col != 'colx']
                    columns_to_plot = [col for col in columns_to_plot if col != 'Event']
                    columns_to_plot = [col for col in columns_to_plot if col not in colu]
                    labels = [str(column) for column in columns_to_plot]
                    

                    df['colx'] = df['colx'].astype(str)
                    df['colx'] = df['colx'].str.split('.').str[0]
                    df = df.loc[:, ~df.columns.duplicated()]

                    plt.figure(figsize=(20, 12))  # Set the figure size
                    fig, ax = plt.subplots() 
                    df.to_csv('prova3.csv', index=False)
                    if checkbox:
                        if 'Limit1' in columns_to_plot:
                            columns_to_plot.remove('Limit1') 
                        if 'Limit2' in columns_to_plot:
                            columns_to_plot.remove('Limit2')
                    try:
                        for column in columns_to_plot:
                            df = df.dropna(subset=column)
                            ax.plot(df['colx'], df[column], label=str(column))

                    except TypeError as e:
                        for i in range(len(columns_to_plot)):
                        
                            x = df['colx']  
                            y = df[columns_to_plot[i]]
                            y = [str(val) for val in y]
                            ax.scatter(x, y, label=labels[i], s=5)

                    if minimum:
                        ax.axhline(y=float(minimum), linestyle='--', label='Reference Limit1', color='black')
                    if maximum:
                        if var_unit3.get()=="x<Limit1 or x>Limit2" or var_unit3.get()=="Limit1<x<Limit2" or var_unit3.get()=="Limit1<=x<=Limit2":
                            ax.axhline(y=float(maximum), linestyle='--', label='Reference Limit2', color='black')
                    ax.xaxis.set_major_locator(plt.MaxNLocator(6))
                    ax.yaxis.set_major_locator(plt.MaxNLocator(6))
                    ax.legend(loc='upper left', bbox_to_anchor=(1.05, 1)) 
                    ax.set_xlabel('Samples')
                    ax.set_ylabel('Values in '+var_unit.get())
                    columns_to_plot = list(columns_to_plot)
                    if checkbox:
                        if 'Limit1' in columns_to_plot:
                            columns_to_plot.remove('Limit1') 
                        if 'Limit2' in columns_to_plot:    
                            columns_to_plot.remove('Limit2')
                    title='Values: '+ ", ".join(str(x) for x in columns_to_plot)
                    ax.set_title(title)
                    fig = ax.figure # Get figure reference 
                    fig.savefig('plot.png', bbox_inches='tight')
                    plt.close(fig)
                    
                    img = Image.open('plot.png')
                    excel_img = ExcelImage(img)
                    # Load the workbook using openpyxl
                    book = openpyxl.load_workbook(output_file)
                    
                    sheet_name = 'Data'+str(currentpage)  # Replace with the actual sheet name
                    sheet = book[sheet_name]
                    # Trova prima colonna vuota
                    empty_col=len(selection['File1']) + len(selection['File2']) + 2

                    # Aggiungi immagine nella colonna vuota, riga 1 
                    sheet.add_image(excel_img, f'{openpyxl.utils.get_column_letter(empty_col)}1')  
                    book.save(output_file)
                    checkp=1
                    print(df.columns)
                    #print(df['Event'])
                    #if df['Event']:
                    if 'Event' in df.columns:
                        df = df.drop(columns='Event')
                    if 'df0_keepevent' in locals() or 'df0_keepevent' in globals():
                        df=pd.concat([df, df0_keepevent], axis=1)           
                    
                    break
                    
            elif not check1:

                fig, ax = plt.subplots()
                if checkbox:
                    if maximum and var_unit3.get()=="Limit1<x<Limit2" or var_unit3.get()=="Limit1<x<Limit2":
                        maxp=pd.Series([float(maximum)]*len(df.iloc[:,0]), name= 'Limit2')
                        df=pd.concat([df, maxp], axis=1)
                    if minimum:
                        minp=pd.Series([float(minimum)]*len(df.iloc[:,0]), name= 'Limit1')
                        df=pd.concat([df, minp], axis=1)
                
                df = df.loc[:, ~df.columns.duplicated()]

                columns_to_plot = df.columns
                
                df.to_csv('prova2.csv', index=False)
                
                columns_to_plot = [col for col in columns_to_plot if col != 'Event']

                check1, format_date, colu=is_date_column2('prova2.csv')
                if check1:
                    columns_to_plot = [col for col in columns_to_plot if col not in colu]

                for column in columns_to_plot:
                    
                    column_dtype=df[column].dtype
                    
                    if column_dtype!='float64':
                        df[column] = df[column].str.replace(',', '.')
                        df[column] = pd.to_numeric(df[column], errors='coerce')
                        df[column] = df[column].apply(lambda x: round(x, 2))
                    
                labels = [str(column) for column in columns_to_plot]

                plt.figure(figsize=(10, 6))  # Set the figure size

                if checkbox:
                    if 'Limit1' in columns_to_plot:
                        columns_to_plot.remove('Limit1') 
                    if 'Limit2' in columns_to_plot:
                        columns_to_plot.remove('Limit2')
                for column in columns_to_plot:
                    df = df.dropna(subset=column)
                    y = df[column]
                    x = range(len(y))
                    ax.plot(x, y, label=str(column))
                
                if minimum:
                    ax.axhline(y=float(minimum), linestyle='--', label='Reference Limit1', color='black')
                if maximum:
                    if var_unit3.get()=="x<Limit1 or x>Limit2" or var_unit3.get()=="Limit1<x<Limit2" or var_unit3.get()=="Limit1<=x<=Limit2":
                        ax.axhline(y=float(maximum), linestyle='--', label='Reference Limit2', color='black')
                # Customize the plot
                ax.xaxis.set_major_locator(plt.MaxNLocator(6))
                ax.yaxis.set_major_locator(plt.MaxNLocator(6))
                ax.legend(loc='upper left', bbox_to_anchor=(1.05, 1))
                ax.set_xlabel('Samples')
                ax.set_ylabel('Values in '+var_unit.get())
                columns_to_plot = list(columns_to_plot)
                if checkbox:
                    if 'Limit1' in columns_to_plot:
                        columns_to_plot.remove('Limit1') 
                    if 'Limit2' in columns_to_plot:
                        columns_to_plot.remove('Limit2')
                title='Values'+ ", ".join(str(x) for x in columns_to_plot)
                ax.set_title(title)
                

                fig = ax.figure # Get figure reference 
                fig.savefig('plot.png', bbox_inches='tight')
                plt.close(fig)
                
                img = Image.open('plot.png')
                excel_img = ExcelImage(img)
                # Load the workbook using openpyxl
                book = openpyxl.load_workbook(output_file)
                
                sheet_name = 'Data'+str(currentpage)  # Replace with the actual sheet name
                sheet = book[sheet_name]
                # Trova prima colonna vuota
                empty_col=len(selection['File1']) + len(selection['File2']) + 2

                # Aggiungi immagine nella colonna vuota, riga 1 
                sheet.add_image(excel_img, f'{openpyxl.utils.get_column_letter(empty_col)}1')  
                book.save(output_file)
                # if os.path.exists('interactive_plot.txt'):
                #     pass
                # else:
                #     print("launching interactive plot")
                #     launch_interactive_plot(df)
                

        if n_files==itera_tion:
                
                loading_label.destroy()
                root.withdraw()
                if os.path.isfile('data.csv'):
                    os.remove('data.csv')
                if os.path.isfile('backup.csv'):
                    os.remove('backup.csv')
                if os.path.isfile('backupdf2.csv'):
                    os.remove('backupdf2.csv')
                if os.path.isfile('backupprova.csv'):    
                    os.remove('backupprova.csv')
                if os.path.isfile('prova2.csv'):
                    os.remove('prova2.csv')         
                if os.path.isfile('output0_modified.csv'):
                    os.remove('output0_modified.csv')
                if os.path.isfile('output1_modified.csv'):
                    os.remove('output1_modified.csv')
                if os.path.isfile('plot.png'):    
                    os.remove('plot.png')
                if os.path.isfile('prova.csv'):    
                    os.remove('prova.csv')
                if os.path.isfile('prova1.csv'):    
                    os.remove('prova1.csv')
                if os.path.isfile('prova3.csv'):    
                    os.remove('prova3.csv')           
                if os.path.isfile('backupdf1.csv'):    
                    os.remove('backupdf1.csv')
                if os.path.isfile('backupdoasjfouieqhfiufho.csv'):    
                    os.remove('backupdoasjfouieqhfiufho.csv')
                if os.path.isfile('output0check.csv'):    
                    os.remove('output0check.csv')
                if os.path.isfile('output1check.csv'):    
                    os.remove('output1check.csv')
                if os.path.isfile('your_additional_output_file.csv'):
                        os.remove('your_additional_output_file.csv')
                messagebox.showinfo("Done", "The analysis have been saved", parent=root)    
                print(enable_plot)
                book = openpyxl.load_workbook(output_file)
                sheet_name = 'Data' + str(currentpage)  # Replace with the actual sheet name.
                sheet = book[sheet_name]

                # Check for duplicate "Event" columns in row 1.
                event_columns = []
                for col in range(1, sheet.max_column + 1):
                    if sheet.cell(row=1, column=col).value == "Event":
                        event_columns.append(col)

                # If exactly two "Event" columns exist, delete the second (last) one.
                if len(event_columns) == 2:
                    col_to_delete = event_columns[-1]
                    sheet.delete_cols(col_to_delete)
                    print(f"Deleted duplicate 'Event' column at index {col_to_delete}")
                book.save(output_file)
                #print(df2['Event'])
                
                if enable_plot==True:
                    if os.path.exists('interactive_plot.txt'):
                        pass
                    else:
                        print("launching interactive plot")
                        if select==2:
                            if 'Event' in df2.columns and( 'df0_keepevent' in locals() or 'df0_keepevent' in globals()):
                                df2 = df2.drop(columns='Event')
                                df2.insert(1, 'Event', df0_keepevent)

                            launch_interactive_plot(root,df0,df2)
                        else:
                            launch_interactive_plot(root, df)
    except Exception as e:
        root=tk.Tk()
        root.withdraw()
        messagebox.showerror("Critical Error", str(e.args))
        loading_label.destroy()
    

    
class ErrorDialog(tk.Toplevel):
    def __init__(self, parent, message, title):
        super().__init__(parent)
        screen_width = parent.winfo_screenwidth()
        screen_height = parent.winfo_screenheight()
        parent.geometry(f"+{screen_width // 2 - parent.winfo_width() // 2}+{screen_height // 2 - parent.winfo_height() // 2}")
        self.title(title)
        self.message = message

        # crea un'etichetta per il messaggio di errore
        label = tk.Label(self, text=self.message)
        label.pack()

        # crea un pulsante per chiudere la finestra di dialogo
        button = tk.Button(self, text="Close", command=self.destroy)
        button.pack()
  
def find_values_out_of_range(column, min_value, max_value):
    try:
        out_of_range_indices = []
        for index, value in enumerate(column):
            if not (min_value <= value <= max_value):
                out_of_range_indices.append(index)
        return out_of_range_indices
    except Exception as e:
        root=tk.Tk()
        root.withdraw()
        print(e)
        error_message = e.args
        messagebox.showerror("Critical Error", str(error_message))



def check_finale(df):
    try:
        df.iloc[:,0] = pd.to_datetime(df.iloc[:,0],format='%H:%M:%S')
        first_sample = df.iloc[:,0].min()
        first_sample = pd.to_datetime(first_sample)
        first_minute_df = df[(df.iloc[:, 0] >= first_sample) & (df.iloc[:, 0] < first_sample + pd.Timedelta(minutes=1))]

        # Conta il numero di campioni nel primo minuto
        num_samples = len(first_minute_df)
        
        # Verifica se il numero di campioni è inferiore a 60
        if num_samples < 60:
            # Calcola il numero di valori 0 da aggiungere
            num_zeros = 60 - num_samples
            
            # Crea un DataFrame con i valori 0 da aggiungere
            zeros_df = pd.DataFrame({column: [0] * num_zeros for column in df.columns[0:]})
            zeros_df.iloc[:, 0] = first_sample
            # Concatena il DataFrame dei valori 0 al DataFrame originale
            df = pd.concat([zeros_df, df], ignore_index=True)
        return df
    except Exception as e:
        root=tk.Tk()
        root.withdraw()
        print(e)
        error_message = e.args
        messagebox.showerror("Critical Error", str(error_message))


# Select user-chosen columns from dataframes
def select_2columns(df1, df2):
    try:
        global j
        global currentpage
        df1= remove_special_characters(df1)
        df2= remove_special_characters(df2)
        selection['File1'] = remove_special_characters_from_list(selection['File1'])
        selection['File2'] = remove_special_characters_from_list(selection['File2'])
        for i, element in enumerate(selection['File1']):
                selection['File1'][i] = element.replace('[', '').replace(']','')
        for i, element in enumerate(selection['File2']):
                selection['File2'][i] = element.replace('[', '').replace(']','')
        df1.columns = df1.columns.str.replace('[', '').str.replace(']', '')
        df2.columns = df2.columns.str.replace('[', '').str.replace(']', '')
        
        selected_df1 = df1[selection['File1']]  # Seleziona solo le colonne desiderate
        selected_df2 = df2[selection['File2']]  # Seleziona solo le colonne desiderate
        selected_df1 = selected_df1.reset_index(drop=True)
        selected_df2 = selected_df2.reset_index(drop=True)

        merged_df = pd.concat([selected_df1, selected_df2], axis=1)
        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl') as writer:
            merged_df.to_excel(writer, sheet_name='Data'+str(j), index=False)
        merged_df.to_csv('backup.csv', index=False)
        
        currentpage=j
        j=j+1
    except Exception as e:
        root=tk.Tk()
        root.withdraw()
        print(e)
        error_message = e.args
        messagebox.showerror("Critical Error", str(error_message))


def select_columns(df1):
    try:
        global j
        global currentpage
        df1= remove_special_characters(df1)
        
        selection['File1'] = remove_special_characters_from_list(selection['File1'])
        for i, element in enumerate(selection['File1']):
                selection['File1'][i] = element.replace('[', '').replace(']','')
                
        df1.columns = df1.columns.str.replace('[', '').str.replace(']', '')
        #df1.columns = df1.columns.str.split('.').str[0]
        # Check if selection['File1'] is a list
        
        # if isinstance(selection['File1'], list):
        #     # If it is a list, then split the elements of the list at the '.' and take the first element
        #     selection['File1'] = [element.split('.')[0] for element in selection['File1']]
        # else:
        #     # If it is not a list, then split the element at the '.' and take the first element
        #     selection['File1'] = selection['File1'].split('.')[0]

        try: 
            selected_df1 = df1[selection['File1']]  # Seleziona solo le colonne desiderate
        except KeyError:
            df1.columns = df1.iloc[0].tolist()
            df1.columns = df1.columns.astype(str)
        
            selected_df1 = df1[selection['File1']]  # Seleziona solo le colonne desiderate

        # Salva il DataFrame nel file CSV esistente con il secondo foglio
        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl') as writer:
            selected_df1.to_excel(writer, sheet_name='Data'+str(j), index=False)
        selected_df1.to_csv('backup.csv', index=False)
        currentpage=j
        j=j+1
    except Exception as e:
        root=tk.Tk()
        root.withdraw()
        print(e)
        error_message = e.args
        messagebox.showerror("Critical Error", str(error_message))