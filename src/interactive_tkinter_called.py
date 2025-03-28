import tkinter as tk
from tkinter import ttk, colorchooser, filedialog, messagebox
from tkinter.simpledialog import askstring
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.ticker import MaxNLocator
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
from PIL import Image
import os
import math
import re
from src.usefull_functions import apply_formulas_to_column, convert_to_relative_time, align_dataframes
from datetime import datetime
from matplotlib.text import Annotation  # For identifying annotation objects

def formulas_to_column(df1, formula, df2=None):
        """
        Calcola una formula personalizzata utilizzando i valori delle colonne di uno o due dataframe.
        Supporta funzioni matematiche senza il prefisso np.
        
        Args:
            df1 (pandas.DataFrame): Il primo dataframe contenente i dati
            formula (str): La formula da calcolare, contenente nomi di colonne
            df2 (pandas.DataFrame, optional): Il secondo dataframe contenente dati aggiuntivi
            
        Returns:
            pandas.Series: Una serie contenente i risultati del calcolo
        """
        import pandas as pd
        import numpy as np
        import re
        
        # Crea copie sicure dei dataframe
        df1_copy = df1.copy()
        
        # Sostituisci le funzioni matematiche comuni senza il prefisso np.
        # Lista delle funzioni matematiche da sostituire
        math_functions = ['sin', 'cos', 'tan', 'exp', 'log', 'sqrt', 'abs']
        
        # Crea un pattern regex per trovare queste funzioni quando non sono precedute da "np."
        pattern = r'(?<!np\.)\b(' + '|'.join(math_functions) + r')\('
        
        # Sostituisci ogni occorrenza con "np.funzione("
        modified_formula = formula
        for match in re.finditer(pattern, formula):
            func_name = match.group(1)
            start_pos = match.start()
            end_pos = match.start() + len(func_name)
            
            # Verifica che non sia parte di un nome di colonna più lungo
            if start_pos > 0 and formula[start_pos-1].isalnum():
                continue
                
            # Sostituisci la funzione con np.funzione
            modified_formula = modified_formula.replace(func_name + '(', 'np.' + func_name + '(', 1)
        
        # Trova tutti i nomi di colonne nella formula
        column_names = set(re.findall(r'[a-zA-Z_][a-zA-Z0-9_]*', modified_formula))
        
        # Parole chiave Python e funzioni matematiche da escludere dalla verifica delle colonne
        excluded_words = ['and', 'or', 'not', 'True', 'False', 'if', 'else', 'for', 'in', 'return', 'None']
        excluded_words.extend(['np'] + math_functions)
        
        # Crea un dizionario locale con le colonne del primo dataframe
        local_dict = {}
        
        # Verifica le colonne nel primo dataframe
        for col in column_names:
            if col not in excluded_words:
                if col in df1_copy.columns:
                    local_dict[col] = df1_copy[col].values
                elif df2 is not None and col in df2.columns:
                    # Se la colonna è nel secondo dataframe, aggiungiamola al dizionario
                    local_dict[col] = df2[col].values
                else:
                    raise ValueError(f"Colonna '{col}' non trovata in nessuno dei dataframe")
        
        # Aggiungi funzioni matematiche comuni
        local_dict.update({
            'np': np,
            'sin': np.sin,
            'cos': np.cos,
            'tan': np.tan,
            'exp': np.exp,
            'log': np.log,
            'sqrt': np.sqrt,
            'abs': np.abs
        })
        
        try:
            # Valuta la formula utilizzando eval (sicuro perché limitato al dizionario locale)
            result = eval(modified_formula, {"__builtins__": {}}, local_dict)
            
            # Converti il risultato in una Serie pandas
            if isinstance(result, np.ndarray):
                return pd.Series(result, index=df1_copy.index)
            else:
                return pd.Series([result] * len(df1_copy), index=df1_copy.index)
        except Exception as e:
            raise ValueError(f"Errore nel calcolo della formula: {str(e)}")

def process_event(checkbox_event, df1, df2=None, options_file='options_event.txt', sec=None):
    if not checkbox_event:
        return df1, df2
    indices_event = []
    with open(options_file, 'r') as f:
        for line in f:
            parts = line.split('@#@')
            if len(parts) > 1:
                event_name = parts[1].strip()
                matching_idx = df1.index[df1['Event'] == event_name]
                if not matching_idx.empty:
                    indices_event.append(int(matching_idx.min()))
    indices_event.sort()
    if len(indices_event) == 1:
        start_event = indices_event[0]
        stop_event = df1.index[-1-int(sec)] if sec else df1.index[-1]
    elif len(indices_event) >= 2:
        start_event = indices_event[0]
        stop_event = indices_event[1]
    else:
        start_event, stop_event = df1.index[0], df1.index[-1]
    df1 = df1.iloc[start_event:stop_event+1]
    start_line = df1.iloc[0, 0]
    stop_line = df1.iloc[-1, 0]
    if df2 is not None:
        try:
            start_idx_df2 = df2.index[df2.iloc[:, 0] == start_line].tolist()[0]
        except IndexError:
            start_idx_df2 = df2.index[0]
        try:
            stop_idx_df2 = df2.index[df2.iloc[:, 0] == stop_line].tolist()[0]
        except IndexError:
            stop_idx_df2 = df2.index[-1]
        df2 = df2.iloc[start_idx_df2:stop_idx_df2+1]
    return df1, df2

def safe_set_message(self, s):
    try:
        self.message.set(s)
    except RuntimeError:
        pass

NavigationToolbar2Tk.set_message = safe_set_message

class FastZoomToolbar2Tk(NavigationToolbar2Tk):
    def __init__(self, canvas, window):
        super().__init__(canvas, window)
        
        # Initialize zoom-related variables
        self._zoom_rect = None
        self._zoom_background = None
        self._zoom_start = None
        self._zoom_start_data = None
        self._zoom_active = False
        self.zoom_mode = None  # 'in' or 'out' or None
        
        # Store original limits
        self.has_stored_limits = False
        self.original_xlim = None
        self.original_ylim = None
        
        # Connect event handlers
        self.canvas.mpl_connect('button_press_event', self._on_press)
        self.canvas.mpl_connect('button_release_event', self._on_release)
        self.canvas.mpl_connect('motion_notify_event', self._on_motion)
        
        # Create zoom buttons
        self.zoom_in_button = tk.Button(self, text="Zoom In", command=self._toggle_zoom_in)
        self.zoom_in_button.pack(side=tk.LEFT, padx=2, pady=2)
        
        self.zoom_out_button = tk.Button(self, text="Zoom Out", command=self._toggle_zoom_out)
        self.zoom_out_button.pack(side=tk.LEFT, padx=2, pady=2)
        # Override home button functionality
        for text, tooltip_text, image_file, callback in self.toolitems:
            if text == 'Home':
                self._buttons[text].config(command=self._home_reset)
                break
    
    def _store_original_limits(self):
        """Store the original limits if not already stored"""
        if not self.has_stored_limits:
            for ax in self.canvas.figure.axes:
                self.original_xlim = ax.get_xlim()
                self.original_ylim = ax.get_ylim()
                self.has_stored_limits = True
                break
    
    def _home_reset(self):
        """Reset view to original limits"""
        if self.has_stored_limits:
            for ax in self.canvas.figure.axes:
                ax.set_xlim(self.original_xlim)
                ax.set_ylim(self.original_ylim)
            self.canvas.draw()
    
    def _toggle_zoom_in(self):
        """Toggle zoom in mode"""
        self._store_original_limits()
        if self.zoom_mode == 'in':
            self._exit_zoom_mode()
        else:
            self.zoom_mode = 'in'
            self._configure_zoom_buttons(self.zoom_in_button)
            
    
    def _toggle_zoom_out(self):
        """Toggle zoom out mode"""
        self._store_original_limits()
        if self.zoom_mode == 'out':
            self._exit_zoom_mode()
        else:
            self.zoom_mode = 'out'
            self._configure_zoom_buttons(self.zoom_out_button)
              
    def _configure_zoom_buttons(self, active_button):
        """Configure zoom button appearance"""
        active_button.config(bg="green", fg="white")
    
    def _exit_zoom_mode(self):
        """Exit zoom mode"""
        self.zoom_mode = None
        self.zoom_in_button.config(bg="SystemButtonFace", fg="black")
        self.zoom_out_button.config(bg="SystemButtonFace", fg="black")
        self.canvas.get_tk_widget().config(cursor="")
    
    def _on_press(self, event):
        """Handle mouse press events"""
        self._store_original_limits()
        
        # Handle zoom click
        if self.zoom_mode and event.inaxes:
            self._perform_zoom_click(event)
            return
        
        # Handle zoom rectangle
        if self.mode == 'zoom' and event.inaxes:
            self._zoom_active = True
            self._zoom_start = (event.x, event.y)
            self._zoom_start_data = (event.xdata, event.ydata)
            
            # Create or update zoom rectangle
            if self._zoom_rect is None:
                self._zoom_rect = plt.Rectangle(
                    (event.xdata, event.ydata), 0, 0,
                    fill=False, color='black', linestyle='--'
                )
                event.inaxes.add_patch(self._zoom_rect)
            else:
                self._zoom_rect.set_visible(True)
                self._zoom_rect.set_xy((event.xdata, event.ydata))
                self._zoom_rect.set_width(0)
                self._zoom_rect.set_height(0)
            
            self.canvas.draw()
            self._zoom_background = self.canvas.copy_from_bbox(event.inaxes.bbox)
    
    def _on_release(self, event):
        """Handle mouse release events"""
        if not self._zoom_active or event.inaxes is None:
            return
        
        self._zoom_active = False
        
        # Apply zoom if rectangle has size
        if self._zoom_rect is not None:
            x0, y0 = self._zoom_rect.get_x(), self._zoom_rect.get_y()
            width, height = self._zoom_rect.get_width(), self._zoom_rect.get_height()
            
            if width > 0 and height > 0:
                event.inaxes.set_xlim(x0, x0 + width)
                event.inaxes.set_ylim(y0, y0 + height)
            
            self._zoom_rect.set_visible(False)
        
        self.canvas.draw()
        self._zoom_start = None
        self._zoom_start_data = None
    
    def _on_motion(self, event):
        """Handle mouse motion events"""
        if (not self._zoom_active or event.inaxes is None or
                self._zoom_rect is None or self._zoom_background is None):
            return
        
        # Update zoom rectangle
        ax = event.inaxes
        self.canvas.restore_region(self._zoom_background)
        
        x0, y0 = self._zoom_start_data
        x1, y1 = event.xdata, event.ydata
        xmin, ymin = min(x0, x1), min(y0, y1)
        width, height = abs(x1 - x0), abs(y1 - y0)
        
        self._zoom_rect.set_xy((xmin, ymin))
        self._zoom_rect.set_width(width)
        self._zoom_rect.set_height(height)
        
        ax.draw_artist(self._zoom_rect)
        self.canvas.blit(ax.bbox)
    
    def _perform_zoom_click(self, event):
        """Perform zoom in/out at click location"""
        ax = event.inaxes
        x_center, y_center = event.xdata, event.ydata
        xlim = ax.get_xlim()
        ylim = ax.get_ylim()
        
        # Calculate zoom factor
        zoom_factor = 0.8 if self.zoom_mode == 'in' else 1.25
        x_range = (xlim[1] - xlim[0]) * zoom_factor
        y_range = (ylim[1] - ylim[0]) * zoom_factor
        
        # Apply zoom
        ax.set_xlim(x_center - x_range / 2, x_center + x_range / 2)
        ax.set_ylim(y_center - y_range / 2, y_center + y_range / 2)
        
        self.canvas.draw()


def _is_time_column(series):
    try:
        pd.to_datetime(series, format="%H:%M:%S")
        return True
    except Exception:
        return False

class PaginatedOptionMenu:
    def __init__(self, master, variable, options, command=None, page_size=10):
        self.master = master
        self.variable = variable
        if not options:
            options = ["Select Event"]
        self.all_options = options
        self.command = command
        self.page_size = page_size
        self.current_page = 0
        self.event_option_var = tk.StringVar(master)
        self.event_option_var.set("Select Event")
        self.option_menu = tk.OptionMenu(
            master,
            self.event_option_var,
            *self.get_current_page_options(),
            command=self.on_select
        )
        self.option_menu.pack(fill=tk.X, padx=5, pady=2)
        self.event_option_var.set("Select Event")
    
    def get_current_page_options(self):
        start = self.current_page * self.page_size
        end = start + self.page_size
        page_options = self.all_options[start:end]
        if self.current_page > 0:
            page_options.insert(0, "< Prev")
        if end < len(self.all_options):
            page_options.append("Next >")
        return page_options
    
    def on_select(self, value):
        if value == "Next >":
            self.current_page += 1
            self.refresh_menu()
            x = self.option_menu.winfo_rootx()
            y = self.option_menu.winfo_rooty() + self.option_menu.winfo_height()
            self.option_menu["menu"].post(x, y)
        elif value == "< Prev":
            self.current_page -= 1
            self.refresh_menu()
            x = self.option_menu.winfo_rootx()
            y = self.option_menu.winfo_rooty() + self.option_menu.winfo_height()
            self.option_menu["menu"].post(x, y)
        else:
            if self.command:
                self.command(value)
            self.event_option_var.set("Select Event")
    
    def refresh_menu(self):
        new_options = self.get_current_page_options()
        menu = self.option_menu["menu"]
        menu.delete(0, "end")
        for option in new_options:
            menu.add_command(
                label=option,
                command=tk._setit(self.event_option_var, option, self.on_select)
            )
        self.event_option_var.set("Select Event")
    
    def update_options(self, new_options):
        self.all_options = new_options
        self.current_page = 0
        self.refresh_menu()


class MovingAveragePopup:
    def __init__(self, parent, plotted_columns, apply_callback):
        """
        parent: the parent widget (InteractivePlotApp instance)
        plotted_columns: list of strings like "DF1: colname" or "DF2: colname" representing the columns currently plotted
        apply_callback: a callback function that will be called with the set of selected columns when the user confirms
        """
        self.parent = parent
        self.plotted_columns = plotted_columns
        self.apply_callback = apply_callback
        self.selected_columns = set()

        self.popup = tk.Toplevel(parent)
        self.popup.title("Select Moving Average Columns")
        self.popup.geometry("300x400")

        # Create an OptionMenu instead of a combobox with an Add button.
        # When an option is selected, the callback (add_column) is invoked automatically.
        self.column_var = tk.StringVar(self.popup)
        if plotted_columns:
            self.column_var.set(plotted_columns[0])
        else:
            self.column_var.set("")
        self.option_menu = tk.OptionMenu(self.popup, self.column_var, *plotted_columns, command=self.add_column)
        self.option_menu.pack(pady=5, fill=tk.X, padx=10)

        # Listbox to display selected columns
        self.listbox = tk.Listbox(self.popup)
        self.listbox.pack(pady=5, fill=tk.BOTH, expand=True, padx=10)
        self.listbox.bind("<Double-Button-1>", self.remove_column)

        # Confirm button to finalize the selection
        self.confirm_button = ttk.Button(self.popup, text="Confirm", command=self.confirm_selection)
        self.confirm_button.pack(pady=5)

    def add_column(self, value):
        if value and value not in self.selected_columns:
            self.selected_columns.add(value)
            self.listbox.insert(tk.END, value)

    def remove_column(self, event):
        selected_index = self.listbox.curselection()
        if selected_index:
            column = self.listbox.get(selected_index[0])
            if column in self.selected_columns:
                self.selected_columns.remove(column)
            self.listbox.delete(selected_index)

    def confirm_selection(self):
        self.apply_callback(self.selected_columns)
        self.popup.destroy()


class InteractivePlotApp(tk.Toplevel):
    """
    Applicazione interattiva per la visualizzazione e l'analisi di dati tramite grafici.
    """
    def __init__(self, parent, df1, df2=None):
        super().__init__(parent)
        self.title("Interactive Plot")
        self.geometry("1000x700")
        print("Initializing Interactive Plot App...")

        self.df1 = df1.copy()
        if df2 is not None:
            self.df2 = df2.copy()
            self.df1 = df1.copy()
            if 'Event' in self.df2.columns:
                self.df2 = self.df2.drop('Event', axis=1)
            if 'Limit1' in self.df2.columns:
                self.df2 = self.df2.drop('Limit1', axis=1)
            if 'Limit2' in self.df2.columns:
                self.df2 = self.df2.drop('Limit2', axis=1)
        else:
            self.df2 = None

        if not _is_time_column(self.df1.iloc[:, 0]):
            x_axis = pd.date_range(start='00:00:00', periods=len(self.df1), freq='1S').strftime('%H:%M:%S')
            self.df1.insert(0, 'Time', x_axis)
        self.time_column = self.df1.columns[0]
        self.fig, self.ax = plt.subplots(figsize=(10, 8))
        self.fig.subplots_adjust(top=0.95, bottom=0.20, right=0.95, left=0.05, hspace=0.2, wspace=0.2)
        self.df1_time = pd.to_datetime(self.df1[self.time_column], format="%H:%M:%S", cache=True)
        if self.df2 is not None:
            if not _is_time_column(self.df2.iloc[:, 0]):
                x_axis = pd.date_range(start='00:00:00', periods=len(self.df2), freq='1S').strftime('%H:%M:%S')
                self.df2.insert(0, 'Time', x_axis)
            self.df2_time_column = self.df2.columns[0]
            self.df2_time = pd.to_datetime(self.df2[self.df2_time_column], format="%H:%M:%S", cache=True)
        else:
            self.df2_time = None

        self.colors_df1 = {}
        
        # Sistema di tracciamento delle colonne plottate
        self.plotted_columns = []  # Lista di tutte le colonne attualmente plottate
        self.plotted_columns_by_plot_selected = []  # Lista delle colonne plottate dal pulsante "Plot Selected"
        self.plotted_difference_columns = []  # Lista delle colonne di differenza plottate
        self.plotted_moving_average_columns = []  # Lista delle colonne di media mobile plottate
        self.plotted_custom_formula_columns = []  # Lista delle colonne di formula personalizzata plottate
        self.colors_df2 = {}
        self.thresholds = []
        self.selected_events = []
        self.custom_events = []
        self.custom_event_plot_times = {}
        self.event_line_labels = {}
        self.data_operation = 'normal'
        self.computed_series = None
        self.common_time = None
        self.computed_label = None
        self.ma_window = None
        self.firstplot = False
        self.initialization_plot = False
        self.xy_data = []
        self.kdtree = None
        self.difference_columns = []  # Will store tuples like ("DF1", "col") for subtracted columns

        self.selected_df1_columns = set()
        self.selected_df2_columns = set()
        # New attribute to store the moving average (MA) columns (as strings like "DF1: colname") 
        self.ma_columns = set()

        self.listbox_tooltip = None
        self.tooltip_after_id = None
        self.tooltip_index = None
        self.tooltip_widget = None
        # Dictionary to keep track of annotations for lines (from pick events)
        self.line_annotations = {}
        # List to store manual annotations so they persist across replotting.
        # (If an annotation is removed, it is also removed from this list.)
        self.manual_annotations = []
        # Temporary storage for event lines (to update legend later)
        self.event_lines = []

        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # LEFT: Plot area.
        left_frame = ttk.Frame(main_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        title_frame = ttk.Frame(left_frame)
        title_frame.pack(fill=tk.X, padx=5, pady=(5, 0))
        ttk.Label(title_frame, text="Chart Title:").pack(side=tk.LEFT, padx=5)
        self.chart_title = tk.StringVar(value="")
        self.chart_title_entry = ttk.Entry(title_frame, textvariable=self.chart_title)
        self.chart_title_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.chart_title_entry.bind("<Return>", lambda e: self.create_plot())

        self.canvas = FigureCanvasTkAgg(self.fig, master=left_frame)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        self.canvas.mpl_connect("button_press_event", self.on_legend_click)

        self.toolbar = FastZoomToolbar2Tk(self.canvas, left_frame)
        self.toolbar.update()
        self.toolbar.pack(side=tk.TOP, fill=tk.X)
        # Create annotation for hover
        self.annot = self.ax.annotate("", xy=(0, 0), xytext=(10, 0),
                                       textcoords="offset points", ha="left", va="center",
                                       bbox=dict(boxstyle="round", fc="w"),
                                       arrowprops=dict(arrowstyle="->"))
        self.annot.set_visible(False)
        self.canvas.mpl_connect("motion_notify_event", self.on_hover)
        # Connect the pick event (for clicking on lines/annotations)
        self.canvas.mpl_connect("pick_event", self.on_pick)

        # RIGHT: Controls.
        right_frame = ttk.Frame(main_frame, width=300)
        right_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        right_frame.pack_propagate(False)
        # DF1 Columns listbox
        df1_frame = ttk.LabelFrame(right_frame, text="DF1 Columns")
        df1_frame.pack(fill=tk.X, pady=2)
        # df1_filter_label = ttk.Label(df1_frame, text="Filter Columns:")
        # df1_filter_label.pack(fill=tk.X, padx=5, pady=2)
        self.df1_filter_var = tk.StringVar()
        self.df1_filter_entry = ttk.Entry(df1_frame, textvariable=self.df1_filter_var)
        self.df1_filter_entry.pack(fill=tk.X, padx=5, pady=2)
        self.df1_filter_entry.bind("<Return>", lambda event: self.populate_df1_listbox())
        self.df1_listbox = tk.Listbox(df1_frame, selectmode=tk.MULTIPLE, exportselection=False, height=3)
        self.df1_listbox.pack(fill=tk.BOTH, padx=5, pady=5)
        self.df1_listbox.bind("<Motion>", self.on_listbox_hover)
        self.df1_listbox.bind("<Leave>", self.on_listbox_leave)
        self.df1_listbox.bind("<<ListboxSelect>>", self.update_df1_selection)
        self.populate_df1_listbox()
        # DF2 Columns listbox
        if self.df2 is not None:
            df2_frame = ttk.LabelFrame(right_frame, text="DF2 Columns")
            df2_frame.pack(fill=tk.X, pady=2)
            # df2_filter_label = ttk.Label(df2_frame, text="Filter Columns:")
            # df2_filter_label.pack(fill=tk.X, padx=5, pady=2)
            self.df2_filter_var = tk.StringVar()
            self.df2_filter_entry = ttk.Entry(df2_frame, textvariable=self.df2_filter_var)
            self.df2_filter_entry.pack(fill=tk.X, padx=5, pady=2)
            self.df2_filter_entry.bind("<Return>", lambda event: self.populate_df2_listbox())
            self.df2_listbox = tk.Listbox(df2_frame, selectmode=tk.MULTIPLE, exportselection=False, height=3)
            self.df2_listbox.pack(fill=tk.BOTH, padx=5, pady=5)
            self.df2_listbox.bind("<Motion>", self.on_listbox_hover)
            self.df2_listbox.bind("<Leave>", self.on_listbox_leave)
            self.df2_listbox.bind("<<ListboxSelect>>", self.update_df2_selection)
            self.populate_df2_listbox()
        else:
            self.df2_listbox = None

        self.plot_btn = ttk.Button(right_frame, text="Plot Selected", command=self.plot_normal)
        self.plot_btn.pack(fill=tk.X, pady=2, padx=5)
        self.reset_view_btn = ttk.Button(right_frame, text="Reset View", command=self.reset_view)
        self.reset_view_btn.pack(fill=tk.X, pady=2, padx=5)
        self.color_btn_df1 = ttk.Button(right_frame, text="Choose DF1 Color", command=self.choose_color_df1)
        self.color_btn_df1.pack(fill=tk.X, pady=2, padx=5)
        if self.df2 is not None:
            self.color_btn_df2 = ttk.Button(right_frame, text="Choose DF2 Color", command=self.choose_color_df2)
            self.color_btn_df2.pack(fill=tk.X, pady=2, padx=5)
        data_ops_frame = ttk.LabelFrame(right_frame, text="Data Operations")
        data_ops_frame.pack(fill=tk.X, pady=2, padx=5)

        self.diff_btn = ttk.Button(data_ops_frame, text="Plot Difference", command=self.plot_difference)
        self.diff_btn.pack(fill=tk.X, padx=5, pady=2)

        self.custom_formula_btn = ttk.Button(data_ops_frame, text="Custom Formula", command=self.open_custom_formula_popup)
        self.custom_formula_btn.pack(fill=tk.X, padx=5, pady=2)

        ma_frame = ttk.Frame(data_ops_frame)
        ma_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(ma_frame, text="MA Window:").pack(side=tk.LEFT, padx=2)
        self.ma_entry = ttk.Entry(ma_frame, width=5)
        self.ma_entry.pack(side=tk.LEFT, padx=2)

        self.ma_btn = ttk.Button(data_ops_frame, text="Plot Moving Average", command=self.plot_moving_average)
        self.ma_btn.pack(fill=tk.X, padx=5, pady=2)

        self.ma_time_btn = ttk.Button(data_ops_frame, text="Plot MA (Time Window)", command=self.plot_moving_average_time)
        self.ma_time_btn.pack(fill=tk.X, padx=5, pady=2)
        # Threshold section
        thresh_frame = ttk.LabelFrame(right_frame, text="Thresholds")
        thresh_frame.pack(fill=tk.X, pady=2, padx=5)
        thresh_inner = ttk.Frame(thresh_frame)
        thresh_inner.pack(fill=tk.X, pady=2, padx=5)
        ttk.Label(thresh_inner, text="Value:").pack(side=tk.LEFT, padx=2)
        self.threshold_entry = ttk.Entry(thresh_inner, width=8)
        self.threshold_entry.pack(side=tk.LEFT, padx=2)
        self.add_thresh_btn = ttk.Button(thresh_inner, text="Add", command=self.add_threshold)
        self.add_thresh_btn.pack(side=tk.RIGHT, padx=2)
        thresh_remove_frame = ttk.Frame(thresh_frame)
        thresh_remove_frame.pack(fill=tk.X, pady=2, padx=5)
        self.rem_thresh_btn = ttk.Button(thresh_remove_frame, text="Remove Last Threshold", command=self.remove_threshold)
        self.rem_thresh_btn.pack(fill=tk.X, padx=2)
        if "Event" in self.df1.columns:
            event_frame = ttk.LabelFrame(right_frame, text="Events")
            event_frame.pack(fill=tk.X, pady=2, padx=5)
            #ttk.Label(event_frame, text="Filter Events:").pack(padx=5, pady=2)
            self.event_filter_var = tk.StringVar()
            self.event_filter_entry = ttk.Entry(event_frame, textvariable=self.event_filter_var)
            self.event_filter_entry.pack(fill=tk.X, padx=5, pady=2)
            self.event_filter_entry.bind("<Return>", lambda event: self.filter_events())
            self.all_events = [(idx, ev) for idx, ev in self.df1["Event"].dropna().items()]
            self.event_option_var = tk.StringVar(event_frame)
            self.event_option_var.set("Select Event")
            formatted_events = [self.format_event(ev_tuple) for ev_tuple in self.all_events]
            if not formatted_events:
                formatted_events = ["Select Event"]
            self.event_menu = PaginatedOptionMenu(event_frame, self.event_option_var, formatted_events,
                                                   command=self.add_event_from_option, page_size=10)
            self.rem_event_btn = ttk.Button(event_frame, text="Remove Last Event", command=self.remove_last_event)
            self.rem_event_btn.pack(fill=tk.X, padx=5, pady=2)
            self.create_event_btn = ttk.Button(event_frame, text="Create Custom Event", command=self.initiate_custom_event)
            self.create_event_btn.pack(fill=tk.X, padx=5, pady=2)
        else:
            self.df1["Event"] = None
            event_frame = ttk.LabelFrame(right_frame, text="Events")
            event_frame.pack(fill=tk.X, pady=2, padx=5)
            self.all_events = []
            ttk.Label(event_frame, text="Filter Events:").pack(padx=5, pady=2)
            self.event_filter_var = tk.StringVar()
            self.event_filter_entry = ttk.Entry(event_frame, textvariable=self.event_filter_var)
            self.event_filter_entry.pack(fill=tk.X, padx=5, pady=2)
            self.event_filter_entry.bind("<KeyRelease>", lambda event: self.filter_events())
            self.event_option_var = tk.StringVar(event_frame)
            self.event_option_var.set("Select Event")
            self.event_menu = PaginatedOptionMenu(event_frame, self.event_option_var, ["Select Event"],
                                                   command=self.add_event_from_option, page_size=10)
            self.rem_event_btn = ttk.Button(event_frame, text="Remove Last Event", command=self.remove_last_event)
            self.rem_event_btn.pack(fill=tk.X, padx=5, pady=2)
            self.create_event_btn = ttk.Button(event_frame, text="Create Custom Event", command=self.initiate_custom_event)
            self.create_event_btn.pack(fill=tk.X, padx=5, pady=2)
        final_frame = ttk.Frame(right_frame)
        final_frame.pack(fill=tk.X, pady=5, padx=5)
        self.new_save_btn = ttk.Button(final_frame, text="Save To Excel", command=self.save_to_excel)
        self.new_save_btn.pack(fill=tk.X, pady=2)
        self.save_btn = ttk.Button(final_frame, text="Append Plot to Excel", command=self.append_plot_to_excel)
        self.save_btn.pack(fill=tk.X, pady=2)
        self.close_btn = ttk.Button(final_frame, text="Close", command=self.destroy)
        self.close_btn.pack(fill=tk.X, pady=2)
        self.plot_normal()

    def open_custom_formula_popup(self):
        popup = tk.Toplevel(self)
        popup.title("Custom Formula")
        popup.geometry("500x350")
        
        # Campo di inserimento della formula
        tk.Label(popup, text="Enter custom formula:").pack(pady=5)
        formula_entry = tk.Entry(popup, width=50)
        formula_entry.pack(pady=5)
        
        # Aggiungi un selettore per scegliere il dataframe
        df_frame = ttk.Frame(popup)
        df_frame.pack(pady=5, fill=tk.X, padx=10)
        
        df_var = tk.StringVar(value="DataFrame 1")
        df_label = tk.Label(df_frame, text="Select DataFrame:")
        df_label.pack(side=tk.LEFT)
        
        # Crea radiobutton per selezionare il dataframe
        df1_radio = tk.Radiobutton(df_frame, text="DataFrame 1", variable=df_var, value="DataFrame 1")
        df1_radio.pack(side=tk.LEFT, padx=5)
        
        # Aggiungi il secondo radiobutton solo se df2 esiste
        if hasattr(self, 'df2') and self.df2 is not None:
            df2_radio = tk.Radiobutton(df_frame, text="DataFrame 2", variable=df_var, value="DataFrame 2")
            df2_radio.pack(side=tk.LEFT, padx=5)
        
        # Frame per il filtro
        filter_frame = ttk.Frame(popup)
        filter_frame.pack(pady=5, fill=tk.X, padx=10)
        
        tk.Label(filter_frame, text="Filter Column:").pack(side=tk.LEFT)
        filter_var = tk.StringVar()
        filter_entry = ttk.Entry(filter_frame, textvariable=filter_var)
        filter_entry.pack(side=tk.LEFT, padx=5)
        
        # Elenco di colonne per df1: si escludono ad esempio il tempo (prima colonna) e "Event"
        base_columns_df1 = [col for col in self.df1.columns if col not in [self.df1.columns[0], "Event"]]
        
        # Elenco di colonne per df2 (se esiste)
        base_columns_df2 = []
        if hasattr(self, 'df2') and self.df2 is not None:
            base_columns_df2 = [col for col in self.df2.columns if col not in [self.df2.columns[0], "Event"]]
        
        # Inizializza con le colonne del df1
        current_base_columns = base_columns_df1.copy()
        filtered_columns = current_base_columns.copy()
        
        # Funzione per inserire la colonna selezionata nella formula
        def insert_column(col_name):
            if col_name != "No match":
                current_text = formula_entry.get()
                pos = formula_entry.index(tk.INSERT)
                new_text = current_text[:pos] + col_name + current_text[pos:]
                formula_entry.delete(0, tk.END)
                formula_entry.insert(0, new_text)
                # Posiziona il cursore dopo la colonna inserita
                formula_entry.icursor(pos + len(col_name))
                # Riporta il focus sull'entry della formula
                formula_entry.focus_set()
        
        # Funzione per aggiornare il menu in base al dataframe selezionato e al filtro
        def update_option_menu(*args):
            # Determina quale dataframe è selezionato
            if df_var.get() == "DataFrame 1":
                current_base_columns = base_columns_df1.copy()
            else:
                current_base_columns = base_columns_df2.copy()
            
            # Applica il filtro
            filter_text = filter_var.get().lower()
            new_options = [col for col in current_base_columns if filter_text in col.lower()]
            
            if not new_options:
                new_options = ["No match"]
            
            # Ricrea il menu con le nuove opzioni
            menu = option_menu["menu"]
            menu.delete(0, "end")
            for col in new_options:
                # Usa una lambda che chiama insert_column quando viene selezionata una voce
                menu.add_command(label=col, command=lambda value=col: insert_column(value))
        
        # Crea l'OptionMenu con le colonne iniziali
        selected_column_var = tk.StringVar()
        if filtered_columns:
            selected_column_var.set(filtered_columns[0])
        else:
            selected_column_var.set("")
        
        option_menu = ttk.OptionMenu(popup, selected_column_var, "", command=insert_column)
        option_menu.pack(pady=5)
        
        # Aggiorna il menu iniziale
        update_option_menu()
        
        # Traccia cambiamenti nel filtro e nella selezione del dataframe
        filter_var.trace("w", update_option_menu)
        df_var.trace("w", update_option_menu)
        
        # Aggiungi esempi di formula
        examples_frame = ttk.Frame(popup)
        examples_frame.pack(pady=5, fill=tk.X, padx=10)
        tk.Label(examples_frame, text="Examples:").pack(anchor="w")
        tk.Label(examples_frame, text="(column1 - column2) * column3").pack(anchor="w")
        tk.Label(examples_frame, text="sin(column1) + sqrt(abs(column2))").pack(anchor="w")
        
        # Pulsante per eseguire il plot della formula
        def on_plot():
            formula = formula_entry.get().strip()
            if not formula:
                messagebox.showerror("Error", "La formula non può essere vuota.")
                return
            popup.destroy()
            self.plot_custom_formula(formula)
        
        plot_btn = ttk.Button(popup, text="Plot Calculation", command=on_plot)
        plot_btn.pack(pady=10)
        
        # Imposta il focus iniziale sul campo della formula
        formula_entry.focus_set()


    def plot_custom_formula(self, formula):
        """
        Funzione per plottare una formula personalizzata.
        Aggiunge la formula calcolata al grafico senza rimuovere le colonne già plottate.
        """
        try:
            # Utilizza la funzione helper apply_formulas_to_column, che ora accetta anche df2
            if hasattr(self, 'df2') and self.df2 is not None:
                computed = formulas_to_column(self.df1, formula, self.df2)
            else:
                computed = formulas_to_column(self.df1, formula)
                
            if computed is None:
                messagebox.showerror("Error", "Errore nell'interpretazione della formula.")
                return
                
            # Salva la serie calcolata
            self.computed_series = computed
            self.computed_label = f"Custom Formula: {formula}"
            
            # Imposta un flag per indicare che c'è una formula personalizzata da plottare
            # ma non cambiare l'operazione corrente (data_operation)
            self.has_custom_formula = True
            
            # Aggiorna le liste di tracciamento
            formula_col_name = self.computed_label
            if formula_col_name not in self.plotted_columns:
                self.plotted_columns.append(formula_col_name)
            if formula_col_name not in self.plotted_custom_formula_columns:
                self.plotted_custom_formula_columns.append(formula_col_name)
            
            # Usa il sistema di plotting esistente
            self.create_plot()
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to compute formula: {e}")


    def on_legend_click(self, event):
        """
        Gestisce il clic sulla legenda per rimuovere elementi specifici dal grafico.
        Aggiorna anche le liste di tracciamento quando un elemento viene rimosso.
        """
        if event.dblclick:  # Rimuovi solo su doppio clic
            # Trova quale testo della legenda è stato cliccato
            clicked_text = None
            for text_obj in self.legend_mapping.keys():
                # Verifica se il clic è avvenuto all'interno del bounding box del testo
                bbox = text_obj.get_window_extent()
                if bbox.contains(event.x, event.y):
                    clicked_text = text_obj
                    break
            
            if clicked_text:
                handle = self.legend_mapping[clicked_text]
                # Ottieni l'etichetta dell'elemento cliccato
                label = clicked_text.get_text()
                
                # Rimuovi l'elemento dalla lista principale di tracciamento
                if label in self.plotted_columns:
                    self.plotted_columns.remove(label)
                
                # Rimuovi anche dalle liste specifiche
                if label in self.plotted_columns_by_plot_selected:
                    self.plotted_columns_by_plot_selected.remove(label)
                if label in self.plotted_difference_columns:
                    self.plotted_difference_columns.remove(label)
                if label in self.plotted_moving_average_columns:
                    self.plotted_moving_average_columns.remove(label)
                if label in self.plotted_custom_formula_columns:
                    self.plotted_custom_formula_columns.remove(label)
                
                # # Nascondi l'elemento
                # handle.set_visible(False)
                
                # Gestisci la rimozione in base al tipo di elemento
                if label.startswith("MA (") and hasattr(self, 'ma_columns'):
                    # Per moving average standard, rimuovi solo quella specifica
                    if "DF1:" in label:
                        col_name = label.split("DF1:")[1].strip()
                        self.ma_columns.discard(f"DF1: {col_name}")
                    elif "DF2:" in label:
                        col_name = label.split("DF2:")[1].strip()
                        self.ma_columns.discard(f"DF2: {col_name}")
                
                elif label.startswith("MA Time (") and hasattr(self, 'ma_columns'):
                    # Per moving average basata sul tempo, rimuovi solo quella specifica
                    if "DF1:" in label:
                        col_name = label.split("DF1:")[1].strip()
                        self.ma_columns.discard(f"DF1: {col_name}")
                    elif "DF2:" in label:
                        col_name = label.split("DF2:")[1].strip()
                        self.ma_columns.discard(f"DF2: {col_name}")
                
                elif label.startswith("Custom Formula:") and hasattr(self, 'has_custom_formula'):
                    # Per formula personalizzata, disattiva solo quella
                    if self.has_custom_formula:
                        self.has_custom_formula = False
                    elif hasattr(self, 'saved_formula_series') and self.saved_formula_series is not None:
                        self.saved_formula_series = None
                        self.saved_formula_label = None
                
                elif label.startswith("Threshold:"):
                    # For thresholds, remove only that specific one
                    threshold_value = float(label.split("Threshold:")[1].strip())
                    if threshold_value in self.thresholds:
                        self.thresholds.remove(threshold_value)
                
                elif label.startswith("Row") and "s):" in label:
                    # For events, remove only that specific one
                    # Extract the event index from the legend
                    if "Row" in label:
                        try:
                            row_idx = int(label.split("Row")[1].split()[0].strip()) - 2
                            for i, (idx, ev_name) in enumerate(self.selected_events):
                                if idx == row_idx:
                                    self.selected_events.pop(i)
                                    # Also remove from custom events if it was one
                                    if (idx, ev_name) in self.custom_events:
                                        self.custom_events.remove((idx, ev_name))
                                    # Remove from custom event plot times if present
                                    if idx in self.custom_event_plot_times:
                                        del self.custom_event_plot_times[idx]
                                    break
                        except (ValueError, IndexError):
                            pass
                
                # Rimuovi l'elemento dalla mappatura della legenda se esiste
                if clicked_text in self.legend_mapping:
                    self.legend_mapping.pop(clicked_text)
                
                # Aggiorna il grafico
                self.create_plot()
                # Hide the element
                handle.set_visible(False)

                # Ricrea la legenda senza l'elemento rimosso
                legend = self.ax.get_legend()
                
                # Verifica che la legenda esista prima di accedere ai suoi metodi
                if legend is not None:
                    # Filtra gli elementi visibili
                    visible_handles = []
                    visible_labels = []
                    
                    for h, t in zip(legend.get_lines(), legend.get_texts()):
                        if h.get_visible() and t != clicked_text:
                            visible_handles.append(h)
                            visible_labels.append(t.get_text())
                
                    # Rimuovi la legenda esistente
                    legend.remove()
                    
                    # Crea una nuova legenda con gli elementi rimanenti
                    if visible_handles:
                        # Determina il numero di colonne dinamicamente
                        if len(visible_labels) <= 5:
                            ncol = 1
                        elif len(visible_labels) <= 10:
                            ncol = 2
                        else:
                            ncol = 3
                        
                        leg = self.ax.legend(visible_handles, visible_labels, loc='upper left', bbox_to_anchor=(-0.05, -0.08),
                                            ncol=ncol, columnspacing=2.0)
                        
                        # Aggiorna la mappatura della legenda
                        self.legend_mapping = {text_obj: handle for text_obj, handle in zip(leg.get_texts(), visible_handles)}
                
                # Aggiorna il grafico
                self.canvas.draw()

    def format_event(self, event_tuple):
        row_idx, ev_name = event_tuple
        common_ref = self.get_common_reference()
        if row_idx in self.custom_event_plot_times:
            ev_sec = self.custom_event_plot_times[row_idx]
        else:
            ev_time = pd.to_datetime(self.df1.loc[row_idx, self.time_column], format="%H:%M:%S")
            ev_sec = (ev_time - common_ref).total_seconds()
        return f"Row {row_idx+2} ({ev_sec:.1f}s): {ev_name}"

    def update_df1_selection(self, event):
        visible = {self.df1_listbox.get(i) for i in range(self.df1_listbox.size())}
        selected_visible = {self.df1_listbox.get(i) for i in self.df1_listbox.curselection()}
        self.selected_df1_columns -= (visible - selected_visible)
        self.selected_df1_columns |= selected_visible

    def update_df2_selection(self, event):
        visible = {self.df2_listbox.get(i) for i in range(self.df2_listbox.size())}
        selected_visible = {self.df2_listbox.get(i) for i in self.df2_listbox.curselection()}
        self.selected_df2_columns -= (visible - selected_visible)
        self.selected_df2_columns |= selected_visible

    def populate_df1_listbox(self):
        self.df1_listbox.delete(0, tk.END)
        filter_text = self.df1_filter_entry.get().strip().lower()
        for col in self.df1.columns:
            if col not in ["Event", self.time_column] and filter_text in col.lower():
                self.df1_listbox.insert(tk.END, col)
        for i in range(self.df1_listbox.size()):
            if self.df1_listbox.get(i) in self.selected_df1_columns:
                self.df1_listbox.select_set(i)

    def populate_df2_listbox(self):
        if self.df2_listbox is None:
            return
        self.df2_listbox.delete(0, tk.END)
        filter_text = self.df2_filter_entry.get().strip().lower()
        for col in self.df2.columns:
            if col not in ["Event", self.df2_time_column] and filter_text in col.lower():
                self.df2_listbox.insert(tk.END, col)
        for i in range(self.df2_listbox.size()):
            if self.df2_listbox.get(i) in self.selected_df2_columns:
                self.df2_listbox.select_set(i)

    def filter_events(self):
        filter_text = self.event_filter_entry.get().strip().lower()
        filtered_events = [self.format_event(ev) for ev in self.all_events if filter_text in self.format_event(ev).lower()]
        self.event_menu.update_options(filtered_events)

    def add_event_from_option(self, selected):
        if not selected or selected == "Select Event":
            return
        m = re.match(r"Row (\d+) \([^)]+\): (.+)", selected)
        if not m:
            messagebox.showerror("Error", "Selected event string format is invalid.")
            return
        row_number = int(m.group(1))
        row_idx = row_number - 2
        ev_name = m.group(2).strip()
        event_tuple = (row_idx, ev_name)
        if event_tuple not in self.selected_events:
            self.selected_events.append(event_tuple)
            self.event_option_var.set("Select Event")
            print('add_event_from_option')
            print(self.firstplot)
            if self.firstplot == True:
                old_xlim = self.ax.get_xlim()
                old_ylim = self.ax.get_ylim()
                self.create_plot()
                self.ax.set_xlim(old_xlim)
                self.ax.set_ylim(old_ylim)
                self.canvas.draw()
            else:
                self.create_plot()

    def get_common_reference(self):
        t1 = self.df1_time
        if self.df2 is not None:
            t2 = self.df2_time
            return min(t1.min(), t2.min())
        else:
            return t1.min()


    def initiate_custom_event(self):
        custom_event = askstring("Custom Event", "Enter custom event name:")
        if custom_event:
            self.custom_event_mode = True
            self.custom_event_name = custom_event
            if custom_event not in [ev[1] for ev in self.custom_events]:
                self.custom_events.append(custom_event)
            #messagebox.showinfo("Custom Event", "Move your mouse over the chart to see the vertical line, then click to add the event.")
            # Connect both the click and the motion event callbacks.
            self.custom_event_cid = self.canvas.mpl_connect("button_press_event", self.on_custom_event_click)
            self.custom_event_motion_cid = self.canvas.mpl_connect("motion_notify_event", self.on_custom_event_motion)


    def on_custom_event_motion(self, event):
        if self.custom_event_mode and event.inaxes == self.ax and event.xdata is not None:
            # Get current y-limits to ensure the line spans the full height.
            ymin, ymax = self.ax.get_ylim()
            if not hasattr(self, 'custom_event_vline') or self.custom_event_vline is None:
                self.custom_event_vline = self.ax.axvline(x=event.xdata, color='gray', linestyle='--', alpha=0.5)
            else:
                # Update both the x data and the y span.
                self.custom_event_vline.set_data([event.xdata, event.xdata], [ymin, ymax])
            self.canvas.draw_idle()


    def on_custom_event_click(self, event):
        if self.custom_event_mode and event.inaxes == self.ax:
            common_ref = self.get_common_reference()
            t_sec = (self.df1_time - common_ref).dt.total_seconds()
            idx = (np.abs(t_sec - event.xdata)).idxmin()
            self.df1.loc[idx, "Event"] = self.custom_event_name
            self.custom_event_plot_times[idx] = event.xdata
            event_tuple = (idx, self.custom_event_name)
            if event_tuple not in self.selected_events:
                self.selected_events.append(event_tuple)
            timestamp = (common_ref + pd.Timedelta(seconds=event.xdata)).strftime('%H:%M:%S')
            messagebox.showinfo("Custom Event", f"Event '{self.custom_event_name}' added at timestamp {timestamp}")
            # Remove the vertical mask if it exists.
            if hasattr(self, 'custom_event_vline') and self.custom_event_vline is not None:
                self.custom_event_vline.remove()
                self.custom_event_vline = None
            # Disconnect both custom event callbacks.
            self.canvas.mpl_disconnect(self.custom_event_motion_cid)
            self.canvas.mpl_disconnect(self.custom_event_cid)
            self.custom_event_mode = False
            # Replot the chart, preserving the current view if needed.
            if self.firstplot:
                old_xlim = self.ax.get_xlim()
                old_ylim = self.ax.get_ylim()
                self.create_plot()
                self.ax.set_xlim(old_xlim)
                self.ax.set_ylim(old_ylim)
                self.canvas.draw()
            else:
                self.create_plot()

    def on_hover(self, event):
        if event.inaxes != self.ax:
            if self.annot.get_visible():
                self.annot.set_visible(False)
                self.canvas.draw_idle()
            return

        for line in self.ax.get_lines():
            contains, attrd = line.contains(event)
            if contains:
                ind = attrd.get('ind', [])
                if len(ind) > 0:
                    index = ind[0]
                    xdata = line.get_xdata()
                    ydata = line.get_ydata()
                    x, y = xdata[index], ydata[index]
                    self.annot.xy = (x, y)
                    self.annot.set_text(line.get_label())
                    self.annot.get_bbox_patch().set_facecolor('w')
                    self.annot.get_bbox_patch().set_alpha(0.8)
                    self.annot.set_visible(True)
                    self.canvas.draw_idle()
                    return
        if self.annot.get_visible():
            self.annot.set_visible(False)
            self.canvas.draw_idle()

    def on_pick(self, event):
        if hasattr(self, "custom_event_mode") and self.custom_event_mode:
            return
        artist = event.artist
        if isinstance(artist, Annotation):
            artist.remove()
            if artist in self.manual_annotations:
                self.manual_annotations.remove(artist)
            for line, ann in list(self.line_annotations.items()):
                if ann == artist:
                    del self.line_annotations[line]
                    break
            self.canvas.draw_idle()
            return
        from matplotlib.lines import Line2D
        if isinstance(artist, Line2D):
            if artist in self.line_annotations:
                return
            x_val = event.mouseevent.xdata if event.mouseevent.xdata is not None else (artist.get_xdata()[0] if len(artist.get_xdata()) > 0 else 0)
            y_val = event.mouseevent.ydata if event.mouseevent.ydata is not None else (artist.get_ydata()[0] if len(artist.get_ydata()) > 0 else 0)
            if artist in self.event_line_labels:
                m = re.match(r"Row \d+ \(([\d\.]+)s\): (.+)", self.event_line_labels[artist])
                if m:
                    elapsed = m.group(1)
                    ev_name = m.group(2)
                    text = f"{ev_name} ({elapsed}s)"
                else:
                    text = self.event_line_labels[artist]
            else:
                text = artist.get_label()
            ann = self.ax.annotate(
                text,
                xy=(x_val, y_val),
                xytext=(5, 5),
                textcoords="offset points",
                bbox=dict(boxstyle="round", fc="w", alpha=0.8),
                arrowprops=dict(arrowstyle="->", color="black"),
                picker=True
            )
            self.line_annotations[artist] = ann
            self.manual_annotations.append(ann)
            self.canvas.draw_idle()

    def reset_view(self):
        self.firstplot = False
        self.selected_df1_columns.clear()
        if self.df2_listbox is not None:
            self.selected_df2_columns.clear()
        self.df1_listbox.selection_clear(0, tk.END)
        if self.df2_listbox is not None:
            self.df2_listbox.selection_clear(0, tk.END)
        self.ax.clear()
        self.canvas.draw()

    def on_listbox_hover(self, event):
        listbox = event.widget
        index = listbox.nearest(event.y)
        if index < 0 or index >= listbox.size():
            self.cancel_tooltip()
            return
        current_text = listbox.get(index)
        if self.tooltip_index == index and self.listbox_tooltip is not None:
            tooltip_width = self.tooltip_label.winfo_reqwidth()
            new_x = event.x_root - tooltip_width - 10
            new_y = event.y_root
            self.listbox_tooltip.wm_geometry(f"+{new_x}+{new_y}")
            return
        else:
            self.cancel_tooltip()
            self.tooltip_index = index
            self.tooltip_widget = listbox
            self.tooltip_after_id = listbox.after(2000, lambda: self.show_tooltip(listbox, current_text, event.x_root, event.y_root))

    def show_tooltip(self, listbox, text, x, y):
        self.listbox_tooltip = tk.Toplevel(listbox)
        self.listbox_tooltip.wm_overrideredirect(True)
        self.listbox_tooltip.attributes("-topmost", True)
        self.tooltip_label = tk.Label(self.listbox_tooltip, text=text, background="white",
                                      foreground="black", relief="solid", borderwidth=1)
        self.tooltip_label.pack()
        self.listbox_tooltip.update_idletasks()
        tooltip_width = self.tooltip_label.winfo_reqwidth()
        new_x = x - tooltip_width - 10
        new_y = y
        self.listbox_tooltip.wm_geometry(f"+{new_x}+{new_y}")

    def cancel_tooltip(self):
        if self.tooltip_after_id is not None:
            if self.tooltip_widget is not None:
                self.tooltip_widget.after_cancel(self.tooltip_after_id)
            self.tooltip_after_id = None
            self.tooltip_index = None
        if self.listbox_tooltip is not None:
            self.listbox_tooltip.destroy()
            self.listbox_tooltip = None
            self.tooltip_index = None

    def on_listbox_leave(self, event):
        self.cancel_tooltip()

    def choose_column_dialog(self, columns, prompt):
        dialog = tk.Toplevel(self)
        dialog.title(prompt)
        tk.Label(dialog, text="Select a column:").pack(padx=10, pady=10)
        selected = tk.StringVar(dialog)
        selected.set(columns[0])
        option_menu = tk.OptionMenu(dialog, selected, *columns)
        option_menu.pack(padx=10, pady=10)
        def on_ok():
            dialog.destroy()
        ok_button = tk.Button(dialog, text="OK", command=on_ok)
        ok_button.pack(padx=10, pady=10)
        dialog.grab_set()
        dialog.wait_window()
        return selected.get()

    def choose_color_df1(self):
        if not self.selected_df1_columns:
            messagebox.showwarning("No Plotted Column", "No columns are currently selected.")
            return
        columns = [col for col in self.df1.columns if col in self.selected_df1_columns]
        if len(columns) == 1:
            column = columns[0]
        else:
            column = self.choose_column_dialog(columns, "Select Column")
            if not column:
                return
        color = colorchooser.askcolor()[1]
        if color:
            self.colors_df1[column] = color
            print('choose color df1')
            print(self.firstplot)
            if self.firstplot == True:
                old_xlim = self.ax.get_xlim()
                old_ylim = self.ax.get_ylim()
                self.create_plot()
                self.ax.set_xlim(old_xlim)
                self.ax.set_ylim(old_ylim)
                self.canvas.draw()
            else:
                self.create_plot()

    def choose_color_df2(self):
        if not self.selected_df2_columns:
            messagebox.showwarning("No Plotted Column", "No DF2 columns are currently selected.")
            return
        columns = [col for col in self.df2.columns if col in self.selected_df2_columns]
        if len(columns) == 1:
            column = columns[0]
        else:
            column = self.choose_column_dialog(columns, "Select DF2 Column")
            if not column:
                return
        color = colorchooser.askcolor()[1]
        if color:
            self.colors_df2[column] = color
            print('choose color df2')
            print(self.firstplot)
            if self.firstplot == True:
                old_xlim = self.ax.get_xlim()
                old_ylim = self.ax.get_ylim()
                self.create_plot()
                self.ax.set_xlim(old_xlim)
                self.ax.set_ylim(old_ylim)
                self.canvas.draw()
            else:
                self.create_plot()

    def add_threshold(self):
        value = self.threshold_entry.get().strip()
        if not value:
            messagebox.showwarning("Input Error", "Threshold entry is empty.")
            return
        try:
            thr = float(value)
            self.thresholds.append(thr)
            self.threshold_entry.delete(0, tk.END)
            print('add_threshold')
            print(self.firstplot)
            if self.firstplot == True:
                old_xlim = self.ax.get_xlim()
                old_ylim = self.ax.get_ylim()
                self.create_plot()
                self.ax.set_xlim(old_xlim)
                self.ax.set_ylim(old_ylim)
                self.canvas.draw()
            else:
                self.create_plot()
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter a valid number for threshold.")

    def remove_threshold(self):
        if self.thresholds:
            self.thresholds.pop()
            print('remove_threshold')
            print(self.firstplot)
            if self.firstplot == True:
                old_xlim = self.ax.get_xlim()
                old_ylim = self.ax.get_ylim()
                self.create_plot()
                self.ax.set_xlim(old_xlim)
                self.ax.set_ylim(old_ylim)
                self.canvas.draw()
            else:
                self.create_plot()
        else:
            messagebox.showinfo("Remove Threshold", "No thresholds to remove.")

    def remove_last_event(self):
        if self.selected_events:
            rem = self.selected_events.pop()
            self.df1.loc[rem[0], "Event"] = None
            if rem in self.custom_events:
                self.custom_events.remove(rem)
            print('remove last event')
            print(self.firstplot)
            if self.firstplot == True:
                old_xlim = self.ax.get_xlim()
                old_ylim = self.ax.get_ylim()
                self.create_plot()
                self.ax.set_xlim(old_xlim)
                self.ax.set_ylim(old_ylim)
                self.canvas.draw()
            else:
                self.create_plot()
        else:
            messagebox.showinfo("Remove Event", "No events to remove.")

    # New method: open a dialog to choose two columns (from DF1/DF2 selections) to subtract
    def choose_difference_columns(self):
        options = []
        for col in self.selected_df1_columns:
            options.append("DF1: " + col)
        if self.df2 is not None and self.selected_df2_columns:
            for col in self.selected_df2_columns:
                options.append("DF2: " + col)
        if len(options) < 2:
            messagebox.showerror("Plot Difference", "Select at least two columns for subtraction.")
            return None, None
        dialog = tk.Toplevel(self)
        dialog.title("Select Columns for Subtraction")
        tk.Label(dialog, text="Select first column to subtract:").pack(padx=10, pady=5)
        first_var = tk.StringVar(dialog)
        first_var.set(options[0])
        first_menu = tk.OptionMenu(dialog, first_var, *options)
        first_menu.pack(padx=10, pady=5)
        tk.Label(dialog, text="Select second column to subtract:").pack(padx=10, pady=5)
        second_var = tk.StringVar(dialog)
        default_second = options[1] if options[1] != options[0] else options[0]
        second_var.set(default_second)
        second_menu = tk.OptionMenu(dialog, second_var, *options)
        second_menu.pack(padx=10, pady=5)
        result = {}
        def on_ok():
            first_selection = first_var.get()
            second_selection = second_var.get()
            if first_selection == second_selection:
                messagebox.showerror("Selection Error", "Please select two different columns.")
                return
            result['first'] = first_selection
            result['second'] = second_selection
            dialog.destroy()
        ok_button = tk.Button(dialog, text="OK", command=on_ok)
        ok_button.pack(pady=10)
        dialog.grab_set()
        self.wait_window(dialog)
        if 'first' in result and 'second' in result:
            return result['first'], result['second']
        else:
            return None, None

    def plot_difference(self):
        """
        Funzione per plottare la differenza tra due colonne selezionate.
        Aggiunge la differenza calcolata al grafico senza rimuovere le colonne già plottate.
        """
        first_sel, second_sel = self.choose_difference_columns()
        if not first_sel or not second_sel:
            self.data_operation = 'normal'
            return
        src1, col1 = first_sel.split(": ", 1)
        src2, col2 = second_sel.split(": ", 1)
        if src1 == "DF1":
            series1 = self.df1[col1]
            t1 = self.df1_time
        else:
            series1 = self.df2[col1]
            t1 = self.df2_time
        if src2 == "DF1":
            series2 = self.df1[col2]
            t2 = self.df1_time
        else:
            series2 = self.df2[col2]
            t2 = self.df2_time
        common_ref = min(t1.min(), t2.min())
        t1_sec = (t1 - common_ref).dt.total_seconds().values
        t2_sec = (t2 - common_ref).dt.total_seconds().values
        common_time = np.union1d(t1_sec, t2_sec)
        interp1 = np.interp(common_time, t1_sec, pd.to_numeric(series1, errors='coerce').values)
        interp2 = np.interp(common_time, t2_sec, pd.to_numeric(series2, errors='coerce').values)
        self.computed_series = interp1 - interp2
        self.common_time = common_time
        self.computed_label = f"Difference: {src1}:{col1} - {src2}:{col2}"
        self.data_operation = 'computed_difference'
        
        # Aggiungi la colonna di differenza alla lista delle colonne plottate
        diff_column_name = self.computed_label
        if diff_column_name not in self.plotted_columns:
            self.plotted_columns.append(diff_column_name)
        
        # Aggiungi alla lista delle colonne di differenza
        if diff_column_name not in self.plotted_difference_columns:
            self.plotted_difference_columns.append(diff_column_name)
        
        # Non rimuoviamo più le colonne originali dalla selezione
        # Questo permette di mantenere le colonne originali nel grafico
        print('plot difference')
        print(self.firstplot)
        # Rimuovi la conservazione dei limiti precedenti per permettere l'adattamento automatico
        self.create_plot()

    # Modified moving average functions to open a popup for column selection.
    def plot_moving_average(self):
        """
        Funzione per plottare la media mobile delle colonne selezionate.
        Aggiunge le medie mobili al grafico senza rimuovere le colonne già plottate.
        """
        selections = []
        selections.extend(["DF1: " + col for col in self.selected_df1_columns])
        if self.df2_listbox is not None:
            selections.extend(["DF2: " + col for col in self.selected_df2_columns])
        if not selections:
            messagebox.showerror("Moving Average", "Select at least one column for moving average.")
            return
        try:
            window = int(self.ma_entry.get().strip())
            if window < 1:
                raise ValueError
            self.ma_window = window
        except ValueError:
            messagebox.showerror("Moving Average", "Enter a valid positive integer for the window.")
            return
        
        # Invece di cambiare data_operation, imposta un flag
        self.has_moving_average = True
        self.ma_type = 'standard'
        
        # Open the popup to select which columns to apply moving average to.
        MovingAveragePopup(self, selections, self.ma_popup_callback)

    def plot_moving_average_time(self):
        """
        Funzione per plottare la media mobile basata sul tempo delle colonne selezionate.
        Aggiunge le medie mobili al grafico senza rimuovere le colonne già plottate.
        """
        selections = []
        selections.extend(["DF1: " + col for col in self.selected_df1_columns])
        if self.df2_listbox is not None:
            selections.extend(["DF2: " + col for col in self.selected_df2_columns])
        if not selections:
            messagebox.showerror("Moving Average (Time)", "Select at least one column for moving average by time window.")
            return
        try:
            window_sec = float(self.ma_entry.get().strip())
            if window_sec <= 0:
                raise ValueError
            self.ma_window = window_sec
        except ValueError:
            messagebox.showerror("Moving Average (Time)", "Enter a valid positive number for the time window (in seconds).")
            return
        
        # Invece di cambiare data_operation, imposta un flag
        self.has_moving_average = True
        self.ma_type = 'time'
        
        MovingAveragePopup(self, selections, self.ma_popup_callback)

    def ma_popup_callback(self, selected_columns):
        """
        Callback dalla finestra MovingAveragePopup: memorizza le colonne selezionate,
        aggiorna le liste di tracciamento e aggiorna il grafico.
        """
        # Memorizza le colonne selezionate per la media mobile
        self.ma_columns = set(selected_columns)
        
        # Aggiorna le liste di tracciamento per le medie mobili
        ma_type_str = "Time" if self.ma_type == 'time' else "Standard"
        window_str = f"{self.ma_window}s" if self.ma_type == 'time' else str(self.ma_window)
        
        # Aggiungi le nuove colonne di media mobile alla lista principale
        for col_name in selected_columns:
            ma_col_name = f"MA {ma_type_str} ({window_str}): {col_name}"
            if ma_col_name not in self.plotted_columns:
                self.plotted_columns.append(ma_col_name)
            if ma_col_name not in self.plotted_moving_average_columns:
                self.plotted_moving_average_columns.append(ma_col_name)
        
        # Aggiorna il grafico
        # Rimuovi la conservazione dei limiti precedenti per permettere l'adattamento automatico
        self.create_plot()

    def save_plot_to_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plot"
        start_row = ws.max_row + 1 if ws["A1"].value is not None else 1
        if self.data_operation in ['computed_difference', 'moving_average', 'moving_average_time', 'custom_formula']:
            if self.data_operation == 'custom_formula' and self.computed_series is not None:
                result_df = pd.DataFrame({self.time_column: self.df1[self.time_column],
                                        self.computed_label: self.computed_series})
            elif self.data_operation == 'computed_difference' and self.computed_series is not None:
                result_df = pd.DataFrame({self.time_column: self.df1[self.time_column], self.computed_label: self.computed_series})
            elif self.data_operation == 'moving_average':
                result_df = pd.DataFrame({self.time_column: self.df1[self.time_column]})
                for col in self.selected_df1_columns:
                    if f"DF1: {col}" in self.ma_columns:
                        ma = self.df1[col].rolling(self.ma_window, min_periods=1).mean()
                        result_df[f"MA ({self.ma_window}): DF1:{col}"] = ma
                    else:
                        result_df[f"DF1: {col}"] = self.df1[col]
                if self.df2 is not None:
                    for col in self.selected_df2_columns:
                        if f"DF2: {col}" in self.ma_columns:
                            ma = self.df2[col].rolling(self.ma_window, min_periods=1).mean()
                            result_df[f"MA ({self.ma_window}): DF2:{col}"] = ma
                        else:
                            result_df[f"DF2: {col}"] = self.df2[col]
            elif self.data_operation == 'moving_average_time':
                result_df = pd.DataFrame({self.time_column: self.df1[self.time_column]})
                for col in self.selected_df1_columns:
                    if f"DF1: {col}" in self.ma_columns:
                        t = self.df1_time
                        series = pd.to_numeric(self.df1[col], errors='coerce')
                        series.index = t
                        window_str = f"{self.ma_window}s"
                        ma = series.rolling(window=window_str, min_periods=1).mean()
                        result_df[f"MA Time ({self.ma_window}s): DF1:{col}"] = ma
                    else:
                        result_df[f"DF1: {col}"] = self.df1[col]
                if self.df2 is not None:
                    for col in self.selected_df2_columns:
                        if f"DF2: {col}" in self.ma_columns:
                            t = self.df2_time
                            series = pd.to_numeric(self.df2[col], errors='coerce')
                            series.index = t
                            window_str = f"{self.ma_window}s"
                            ma = series.rolling(window=window_str, min_periods=1).mean()
                            result_df[f"MA Time ({self.ma_window}s): DF2:{col}"] = ma
                        else:
                            result_df[f"DF2: {col}"] = self.df2[col]
            if "Event" in self.df1.columns:
                result_df["Event"] = self.df1["Event"]
            for c_idx, header in enumerate(result_df.columns, start=1):
                ws.cell(row=start_row, column=c_idx, value=header)
            for r_idx, row in enumerate(result_df.itertuples(index=False), start=start_row+1):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            start_row = start_row + result_df.shape[0] + 2
        else:
            df_to_save = pd.DataFrame({self.time_column: pd.to_datetime(self.df1[self.time_column], format="%H:%M:%S").dt.strftime('%H:%M:%S')})
            for col in self.selected_df1_columns:
                df_to_save[f"DF1: {col}"] = self.df1[col]
            if self.df2 is not None:
                for col in self.selected_df2_columns:
                    df_to_save[f"DF2: {col}"] = self.df2[col]
            if "Event" in self.df1.columns:
                df_to_save["Event"] = self.df1["Event"]
            for c_idx, header in enumerate(df_to_save.columns, start=1):
                ws.cell(row=1, column=c_idx, value=header)
            for r_idx, row in enumerate(df_to_save.itertuples(index=False), start=2):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
        headers = [cell.value for cell in ws[1]]
        if "Event" in headers:
            event_col_index = headers.index("Event") + 1
            ws.delete_cols(event_col_index)
        new_event_col = ws.max_column + 1
        ws.cell(row=1, column=new_event_col, value="Event")
        num_rows = ws.max_row
        for r in range(2, num_rows+1):
            if r-2 < len(self.df1):
                event_val = self.df1.iloc[r-2].get("Event", None)
            else:
                event_val = None
            ws.cell(row=r, column=new_event_col, value=event_val)
        orig_size = self.fig.get_size_inches()
        new_size = orig_size * (4/7)
        self.fig.set_size_inches(new_size)
        self.ax.set_title(self.chart_title_entry.get(), pad=15)
        buf = BytesIO()
        self.fig.savefig(buf, format='png', bbox_inches="tight", dpi=150)
        buf.seek(0)
        excel_img = ExcelImage(buf)
        self.fig.set_size_inches(orig_size)
        img_cell = f"{openpyxl.utils.get_column_letter(ws.max_column + 2)}1"
        ws.add_image(excel_img, img_cell)
        try:
            wb.save(file_path)
            messagebox.showinfo("Saved", f"Plot (and computed data if applicable) saved successfully to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save to Excel file: {e}")
        finally:
            buf.close()

    def save_to_excel(self):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"RapidAnalysis_{timestamp}.xlsx"
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")],
                                                 initialfile=default_filename)
        if not file_path:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data0"
        if self.data_operation in ['computed_difference', 'moving_average', 'moving_average_time']:
            if self.data_operation == 'computed_difference' and self.computed_series is not None:
                data_df = pd.DataFrame({self.time_column: self.df1[self.time_column], self.computed_label: self.computed_series})
            elif self.data_operation == 'moving_average':
                data_df = pd.DataFrame({self.time_column: self.df1[self.time_column]})
                for col in self.selected_df1_columns:
                    if f"DF1: {col}" in self.ma_columns:
                        ma = self.df1[col].rolling(self.ma_window, min_periods=1).mean()
                        data_df[f"MA ({self.ma_window}): DF1:{col}"] = ma
                    else:
                        data_df[f"DF1: {col}"] = self.df1[col]
                if self.df2 is not None:
                    for col in self.selected_df2_columns:
                        if f"DF2: {col}" in self.ma_columns:
                            ma = self.df2[col].rolling(self.ma_window, min_periods=1).mean()
                            data_df[f"MA ({self.ma_window}): DF2:{col}"] = ma
                        else:
                            data_df[f"DF2: {col}"] = self.df2[col]
            elif self.data_operation == 'moving_average_time':
                data_df = pd.DataFrame({self.time_column: self.df1[self.time_column]})
                for col in self.selected_df1_columns:
                    if f"DF1: {col}" in self.ma_columns:
                        t = self.df1_time
                        series = self.df1[col].copy()
                        series.index = t
                        window_str = f"{self.ma_window}s"
                        ma = series.rolling(window=window_str, min_periods=1).mean()
                        data_df[f"MA Time ({self.ma_window}s): DF1:{col}"] = ma
                    else:
                        data_df[f"DF1: {col}"] = self.df1[col]
                if self.df2 is not None:
                    for col in self.selected_df2_columns:
                        if f"DF2: {col}" in self.ma_columns:
                            t = self.df2_time
                            series = self.df2[col].copy()
                            series.index = t
                            window_str = f"{self.ma_window}s"
                            ma = series.rolling(window=window_str, min_periods=1).mean()
                            data_df[f"MA Time ({self.ma_window}s): DF2:{col}"] = ma
                        else:
                            data_df[f"DF2: {col}"] = self.df2[col]
            if "Event" in self.df1.columns:
                data_df["Event"] = self.df1["Event"]
        else:
            data_df = pd.DataFrame({self.time_column: pd.to_datetime(self.df1[self.time_column], format="%H:%M:%S").dt.strftime('%H:%M:%S')})
            for col in self.selected_df1_columns:
                data_df[f"DF1: {col}"] = self.df1[col]
            if self.df2 is not None:
                for col in self.selected_df2_columns:
                    data_df[f"DF2: {col}"] = self.df2[col]
            if "Event" in self.df1.columns:
                data_df["Event"] = self.df1["Event"]
        for c_idx, header in enumerate(data_df.columns, start=1):
            ws.cell(row=1, column=c_idx, value=header)
        for r_idx, row in enumerate(data_df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        headers = [cell.value for cell in ws[1]]
        if "Event" in headers:
            event_col_index = headers.index("Event") + 1
            ws.delete_cols(event_col_index)
        new_event_col = ws.max_column + 1
        ws.cell(row=1, column=new_event_col, value="Event")
        num_rows = ws.max_row
        for r in range(2, num_rows+1):
            if r-2 < len(self.df1):
                event_val = self.df1.iloc[r-2].get("Event", None)
            else:
                event_val = None
            ws.cell(row=r, column=new_event_col, value=event_val)
        orig_size = self.fig.get_size_inches()
        new_size = orig_size * (4/7)
        self.fig.set_size_inches(new_size)
        self.ax.set_title(self.chart_title_entry.get(), pad=15)
        buf = BytesIO()
        self.fig.savefig(buf, format='png', bbox_inches="tight", dpi=150)
        buf.seek(0)
        excel_img = ExcelImage(buf)
        self.fig.set_size_inches(orig_size)
        img_cell = f"{openpyxl.utils.get_column_letter(ws.max_column + 2)}1"
        ws.add_image(excel_img, img_cell)
        try:
            wb.save(file_path)
            messagebox.showinfo("Saved", f"Plot and data saved successfully to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save to Excel file: {e}")
        finally:
            buf.close()

    def append_plot_to_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return
        try:
            wb = openpyxl.load_workbook(file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Cannot open Excel file: {e}")
            return
        sheet_dialog = tk.Toplevel(self)
        sheet_dialog.title("Select or Create Sheet")
        sheet_dialog.geometry("350x200")
        content_frame = ttk.Frame(sheet_dialog)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        confirm_frame = ttk.Frame(sheet_dialog)
        confirm_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)
        sheet_var = tk.StringVar()
        new_sheet_var = tk.StringVar()
        select_frame = ttk.Frame(content_frame)
        select_frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(select_frame, text="Select existing sheet:").pack(pady=5)
        sheet_dropdown = ttk.Combobox(select_frame, textvariable=sheet_var, values=wb.sheetnames, state="readonly")
        sheet_dropdown.pack(pady=5)
        sheet_dropdown.set(wb.sheetnames[0])
        add_page_btn = ttk.Button(select_frame, text="Add New Page", command=lambda: switch_to_new_sheet())
        add_page_btn.pack(pady=5)
        new_sheet_frame = ttk.Frame(content_frame)
        back_btn = ttk.Button(new_sheet_frame, text="← Back to Selection", command=lambda: switch_to_existing_sheet())
        back_btn.pack(pady=5, anchor="w")
        ttk.Label(new_sheet_frame, text="Enter new sheet name:").pack(pady=5)
        new_sheet_entry = ttk.Entry(new_sheet_frame, textvariable=new_sheet_var)
        new_sheet_entry.pack(pady=5)
        def switch_to_new_sheet():
            select_frame.pack_forget()
            new_sheet_frame.pack(fill=tk.BOTH, expand=True)
        def switch_to_existing_sheet():
            new_sheet_frame.pack_forget()
            select_frame.pack(fill=tk.BOTH, expand=True)
        def on_confirm():
            if new_sheet_frame.winfo_ismapped():
                new_sheet_name = new_sheet_var.get().strip()
                if not new_sheet_name:
                    messagebox.showerror("Error", "Please enter a sheet name.")
                    return
                if new_sheet_name in wb.sheetnames:
                    messagebox.showerror("Error", "Sheet name already exists. Choose another name.")
                    return
                ws = wb.create_sheet(title=new_sheet_name)
                save_full_data = True
            else:
                selected_sheet = sheet_var.get().strip()
                ws = wb[selected_sheet]
                save_full_data = False
            sheet_dialog.destroy()
            max_col = ws.max_column
            empty_col = max_col + 2
            formatted_time = pd.to_datetime(self.df1[self.time_column], format="%H:%M:%S").dt.strftime('%H:%M:%S')
            if save_full_data:
                df_to_save = pd.DataFrame({self.time_column: formatted_time})
                for col in self.selected_df1_columns:
                    df_to_save[f"DF1: {col}"] = self.df1[col]
                if self.df2 is not None:
                    for col in self.selected_df2_columns:
                        df_to_save[f"DF2: {col}"] = self.df2[col]
                if "Event" in self.df1.columns:
                    df_to_save["Event"] = self.df1["Event"]
                for c_idx, header in enumerate(df_to_save.columns, start=1):
                    ws.cell(row=1, column=c_idx, value=header)
                for r_idx, row in enumerate(df_to_save.itertuples(index=False), start=2):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
            else:
                result_df = pd.DataFrame()
                if self.data_operation == 'computed_difference' and self.computed_series is not None:
                    result_df[self.computed_label] = self.computed_series
                elif self.data_operation == 'moving_average':
                    for col in self.selected_df1_columns:
                        if f"DF1: {col}" in self.ma_columns:
                            ma = self.df1[col].rolling(self.ma_window, min_periods=1).mean()
                            result_df[f"MA ({self.ma_window}): DF1:{col}"] = ma
                    for col in self.selected_df2_columns:
                        if f"DF2: {col}" in self.ma_columns:
                            ma = self.df2[col].rolling(self.ma_window, min_periods=1).mean()
                            result_df[f"MA ({self.ma_window}): DF2:{col}"] = ma
                elif self.data_operation == 'moving_average_time':
                    for col in self.selected_df1_columns:
                        if f"DF1: {col}" in self.ma_columns:
                            t = self.df1_time
                            series = self.df1[col].copy()
                            series.index = t
                            window_str = f"{self.ma_window}s"
                            ma = series.rolling(window=window_str, min_periods=1).mean()
                            result_df[f"MA Time ({self.ma_window}s): DF1:{col}"] = ma
                start_row = 1
                for c_idx, header in enumerate(result_df.columns, start=1):
                    ws.cell(row=start_row, column=empty_col + c_idx, value=header)
                for r_idx, row in enumerate(result_df.itertuples(index=False), start=start_row+1):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=empty_col + c_idx, value=value)
            headers = [cell.value for cell in ws[1]]
            if "Event" in headers:
                event_col_index = headers.index("Event") + 1
                ws.delete_cols(event_col_index)
            new_event_col = ws.max_column + 1
            ws.cell(row=1, column=new_event_col, value="Event")
            num_rows = ws.max_row
            for r in range(2, num_rows+1):
                if r-2 < len(self.df1):
                    event_val = self.df1.iloc[r-2].get("Event", None)
                else:
                    event_val = None
                ws.cell(row=r, column=new_event_col, value=event_val)
            orig_size = self.fig.get_size_inches()
            new_size = orig_size * (4/7)
            self.fig.set_size_inches(new_size)
            self.ax.set_title(self.chart_title_entry.get(), pad=15)
            buf = BytesIO()
            self.fig.savefig(buf, format='png', bbox_inches="tight", dpi=150)
            buf.seek(0)
            excel_img = ExcelImage(buf)
            self.fig.set_size_inches(orig_size)
            img_cell = f"{openpyxl.utils.get_column_letter(ws.max_column + 2)}1"
            ws.add_image(excel_img, img_cell)
            try:
                wb.save(file_path)
                messagebox.showinfo("Saved", f"Plot (and computed data if applicable) saved successfully to {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to append to Excel file: {e}")
            finally:
                buf.close()
        confirm_button = ttk.Button(confirm_frame, text="Confirm", command=on_confirm)
        confirm_button.pack(pady=10)
    
    def reset_moving_average(self):
        # Metodo per resettare le moving average quando necessario
        if hasattr(self, 'has_moving_average'):
            self.has_moving_average = False
        if hasattr(self, 'ma_columns'):
            self.ma_columns = set()
        if hasattr(self, 'ma_type'):
            self.ma_type = None
        self.create_plot()


    def plot_normal(self):
        """
        Funzione per plottare le colonne selezionate nelle listbox.
        Mantiene le colonne già plottate da altre funzioni e aggiorna solo quelle
        gestite dal pulsante "Plot Selected".
        """
        self.data_operation = 'normal'
        
        # Ottieni le colonne attualmente selezionate
        current_selected_columns = []
        for col in self.df1.columns:
            if col in self.selected_df1_columns:
                current_selected_columns.append(f"DF1: {col}")
        
        if self.df2 is not None:
            for col in self.df2.columns:
                if col in self.selected_df2_columns:
                    current_selected_columns.append(f"DF2: {col}")
        
        # Determina quali colonne rimuovere (quelle che erano nella lista precedente ma non in quella attuale)
        columns_to_remove = [col for col in self.plotted_columns_by_plot_selected if col not in current_selected_columns]
        
        # Rimuovi le colonne che non sono più selezionate dalla lista principale
        for col in columns_to_remove:
            if col in self.plotted_columns:
                self.plotted_columns.remove(col)
        
        # Aggiorna la lista delle colonne plottate dal pulsante "Plot Selected"
        self.plotted_columns_by_plot_selected = current_selected_columns.copy()
        
        # Aggiungi le nuove colonne selezionate alla lista principale se non ci sono già
        for col in current_selected_columns:
            if col not in self.plotted_columns:
                self.plotted_columns.append(col)
        
        print('plot_normal')
        print(self.firstplot)
        # Rimuovi la conservazione dei limiti precedenti per permettere l'adattamento automatico
        self.create_plot()



    def create_plot(self):
        """
        Funzione centrale per creare il grafico.
        Utilizza le liste di tracciamento per determinare quali colonne plottare.
        Adatta automaticamente i limiti del grafico e resetta la vista se non ci sono colonne.
        """
        # --- Preserve manual annotations from pick events --- 
        saved_manual_annotations = []
        for ann in self.manual_annotations:
            saved_manual_annotations.append({
                'text': ann.get_text(),
                'xy': ann.xy,
                'xytext': ann.get_position()
            })
        self.manual_annotations = []
        self.ax.clear()
        self.event_line_labels = {}
        self.xy_data = []
        
        # Verifica se non ci sono colonne plottate e in tal caso esegui un reset della vista
        if not self.plotted_columns:
            self.reset_view()
            return
            
        common_ref = self.get_common_reference()
        
        # Set up used colors tracking
        used_colors = set()
        default_colors = plt.rcParams['axes.prop_cycle'].by_key()['color']
        
        def get_unique_color(used_colors, preferred=None):
            if preferred and preferred not in used_colors:
                return preferred
            for color in default_colors:
                if color not in used_colors:
                    return color
            import random
            while True:
                color = "#%06x" % random.randint(0, 0xFFFFFF)
                if color not in used_colors:
                    return color
        
        # Plotta la formula personalizzata se presente nella lista di tracciamento
        if hasattr(self, 'has_custom_formula') and self.has_custom_formula and hasattr(self, 'computed_series') and self.computed_series is not None:
            t_sec = (self.df1_time - common_ref).dt.total_seconds().values
            color = get_unique_color(used_colors, preferred='blue')
            used_colors.add(color)
            self.ax.plot(t_sec, self.computed_series, label=self.computed_label, color=color, picker=5)
            self.xy_data.extend(list(zip(t_sec, self.computed_series)))
        # Plotta la formula salvata se esiste
        elif hasattr(self, 'saved_formula_series') and self.saved_formula_series is not None:
            t_sec = (self.df1_time - common_ref).dt.total_seconds().values
            color = get_unique_color(used_colors, preferred='blue')
            used_colors.add(color)
            self.ax.plot(t_sec, self.saved_formula_series, label=self.saved_formula_label, color=color, picker=5)
            self.xy_data.extend(list(zip(t_sec, self.saved_formula_series)))
        
        # Plotta la differenza calcolata se presente nella lista di tracciamento
        if self.data_operation == 'computed_difference' and hasattr(self, 'computed_series') and self.computed_series is not None:
            comp_color = get_unique_color(used_colors, preferred='purple')
            used_colors.add(comp_color)
            self.ax.plot(self.common_time, self.computed_series,
                        label=self.computed_label, color=comp_color, picker=5)
            self.xy_data.extend(list(zip(self.common_time, self.computed_series)))
        
        # Plotta le colonne normali di DF1 che sono nella lista di tracciamento
        df1_columns_to_plot = []
        for col in self.df1.columns:
            col_name = f"DF1: {col}"
            if col_name in self.plotted_columns:
                df1_columns_to_plot.append(col)
        
        for col in df1_columns_to_plot:
            try:
                t = self.df1_time
                t_sec = (t - common_ref).dt.total_seconds().values
                y_vals = pd.to_numeric(self.df1[col], errors='coerce').values
                y_vals = np.nan_to_num(y_vals, nan=0.0)
                candidate = self.colors_df1.get(col)
                if candidate is None or candidate in used_colors:
                    candidate = get_unique_color(used_colors)
                    self.colors_df1[col] = candidate
                used_colors.add(candidate)
                self.ax.plot(t_sec, y_vals, label=f"DF1: {col}", color=candidate, picker=5)
                self.xy_data.extend(list(zip(t_sec, y_vals)))
            except Exception as e:
                messagebox.showerror("Plot Error", f"Column '{col}' (DF1) could not be plotted: {e}")
                continue
        
        # Plotta le colonne normali di DF2 che sono nella lista di tracciamento
        if self.df2 is not None:
            df2_columns_to_plot = []
            for col in self.df2.columns:
                col_name = f"DF2: {col}"
                if col_name in self.plotted_columns:
                    df2_columns_to_plot.append(col)
            
            for col in df2_columns_to_plot:
                try:
                    t = self.df2_time
                    t_sec = (t - common_ref).dt.total_seconds().values
                    y_vals = pd.to_numeric(self.df2[col], errors='coerce').values
                    y_vals = np.nan_to_num(y_vals, nan=0.0)
                    candidate = self.colors_df2.get(col)
                    if candidate is None or candidate in used_colors:
                        candidate = get_unique_color(used_colors)
                        self.colors_df2[col] = candidate
                    used_colors.add(candidate)
                    self.ax.plot(t_sec, y_vals, label=f"DF2: {col}", color=candidate, picker=5)
                    self.xy_data.extend(list(zip(t_sec, y_vals)))
                except Exception as e:
                    messagebox.showerror("Plot Error", f"Column '{col}' (DF2) could not be plotted: {e}")
                    continue
        
        # Plotta le moving average che sono nella lista di tracciamento
        if hasattr(self, 'has_moving_average') and self.has_moving_average and hasattr(self, 'ma_columns') and self.ma_columns:
            # Ottieni le colonne di DF1 e DF2 che hanno una media mobile
            ma_df1_columns = []
            ma_df2_columns = []
            
            for col_name in self.plotted_moving_average_columns:
                if "DF1:" in col_name and col_name.split("DF1:")[1].strip() in self.df1.columns:
                    col = col_name.split("DF1:")[1].strip()
                    if f"DF1: {col}" in self.ma_columns:
                        ma_df1_columns.append(col)
                elif "DF2:" in col_name and self.df2 is not None and col_name.split("DF2:")[1].strip() in self.df2.columns:
                    col = col_name.split("DF2:")[1].strip()
                    if f"DF2: {col}" in self.ma_columns:
                        ma_df2_columns.append(col)
            
            if hasattr(self, 'ma_type') and self.ma_type == 'standard':
                # Moving average standard
                for col in ma_df1_columns:
                    try:
                        t = self.df1_time
                        t_sec = (t - common_ref).dt.total_seconds().values
                        y_vals = pd.to_numeric(self.df1[col], errors='coerce').rolling(self.ma_window, min_periods=1).mean().values
                        label = f"MA ({self.ma_window}): DF1:{col}"
                        candidate = get_unique_color(used_colors)
                        used_colors.add(candidate)
                        self.ax.plot(t_sec, y_vals, label=label, color=candidate, picker=5)
                        self.xy_data.extend(list(zip(t_sec, y_vals)))
                    except Exception as e:
                        messagebox.showerror("Plot Error", f"Moving average for column '{col}' (DF1) could not be plotted: {e}")
                
                if self.df2 is not None:
                    for col in ma_df2_columns:
                        try:
                            t = self.df2_time
                            t_sec = (t - common_ref).dt.total_seconds().values
                            y_vals = pd.to_numeric(self.df2[col], errors='coerce').rolling(self.ma_window, min_periods=1).mean().values
                            label = f"MA ({self.ma_window}): DF2:{col}"
                            candidate = get_unique_color(used_colors)
                            used_colors.add(candidate)
                            self.ax.plot(t_sec, y_vals, label=label, color=candidate, picker=5)
                            self.xy_data.extend(list(zip(t_sec, y_vals)))
                        except Exception as e:
                            messagebox.showerror("Plot Error", f"Moving average for column '{col}' (DF2) could not be plotted: {e}")
            
            elif hasattr(self, 'ma_type') and self.ma_type == 'time':
                # Moving average basata sul tempo
                for col in ma_df1_columns:
                    try:
                        t = self.df1_time
                        t_sec = (t - common_ref).dt.total_seconds().values
                        series = pd.to_numeric(self.df1[col], errors='coerce')
                        series.index = t
                        window_str = f"{self.ma_window}s"
                        y_vals = series.rolling(window=window_str, min_periods=1).mean().values
                        label = f"MA Time ({self.ma_window}s): DF1:{col}"
                        candidate = get_unique_color(used_colors)
                        used_colors.add(candidate)
                        self.ax.plot(t_sec, y_vals, label=label, color=candidate, picker=5)
                        self.xy_data.extend(list(zip(t_sec, y_vals)))
                    except Exception as e:
                        messagebox.showerror("Plot Error", f"Time-based moving average for column '{col}' (DF1) failed: {e}")
                
                if self.df2 is not None:
                    for col in ma_df2_columns:
                            try:
                                t = self.df2_time
                                t_sec = (t - common_ref).dt.total_seconds().values
                                series = pd.to_numeric(self.df2[col], errors='coerce')
                                series.index = t
                                window_str = f"{self.ma_window}s"
                                y_vals = series.rolling(window=window_str, min_periods=1).mean().values
                                label = f"MA Time ({self.ma_window}s): DF2:{col}"
                                candidate = get_unique_color(used_colors)
                                used_colors.add(candidate)
                                self.ax.plot(t_sec, y_vals, label=label, color=candidate, picker=5)
                                self.xy_data.extend(list(zip(t_sec, y_vals)))
                            except Exception as e:
                                messagebox.showerror("Plot Error", f"Time-based moving average for column '{col}' (DF2) failed: {e}")
        
        # Plotta le soglie
        for thr in self.thresholds:
            thr_color = get_unique_color(used_colors, preferred='red')
            used_colors.add(thr_color)
            self.ax.axhline(y=thr, color=thr_color, linestyle='dashed', label=f"Threshold: {thr}", picker=5)

        # Plotta gli eventi
        if "Event" in self.df1.columns and self.selected_events:
            for event_info in self.selected_events:
                row_idx, ev_name = event_info
                event_color = get_unique_color(used_colors)
                used_colors.add(event_color)
                if row_idx in self.custom_event_plot_times:
                    ev_sec = self.custom_event_plot_times[row_idx]
                else:
                    try:
                        row_time = pd.to_datetime(self.df1.loc[row_idx, self.time_column], format="%H:%M:%S")
                        ev_sec = (row_time - common_ref).total_seconds()
                    except Exception:
                        continue
                short_label = f"({ev_sec:.1f}s): {ev_name}"
                full_label = f"Row {row_idx+2} ({ev_sec:.1f}s): {ev_name}"
                line = self.ax.axvline(x=ev_sec, color=event_color, linestyle='dotted', picker=5, label=short_label)
                self.event_lines.append((line, full_label))
                self.xy_data.append((ev_sec, 0))
                self.event_line_labels[line] = short_label
        
        # Il resto della funzione rimane invariato
        handles, labels = self.ax.get_legend_handles_labels()

        # Ensure event lines keep their full label
        if self.event_lines:
            for line, full_label in self.event_lines:
                for j, handle in enumerate(handles):
                    if handle == line:
                        labels[j] = full_label
                        break
            self.event_lines = []

        # Reverse order for better legend display
        if handles:
            handles = handles[::-1]
            labels = labels[::-1]

            # Limit legend size to 15 items (adjustable)
            max_items = 15  # 3 columns x 5 rows max
            if len(labels) > max_items:
                handles = handles[:max_items]
                labels = labels[:max_items]

            # Determine number of columns dynamically
            if len(labels) <= 5:
                ncol = 1
            elif len(labels) <= 10:
                ncol = 2
            else:
                ncol = 3

            # Create legend in upper-left with adjusted spacing
            leg = self.ax.legend(handles, labels, loc='upper left', bbox_to_anchor=(-0.05, -0.08),
                                ncol=ncol, columnspacing=2.0)

            # Update legend mapping for interactive removal on double-click
            self.legend_mapping = {text_obj: handle for text_obj, handle in zip(leg.get_texts(), handles)}

        self.ax.set_xlabel("Elapsed Time [s]")
        self.ax.set_ylabel("Values")
        self.ax.ticklabel_format(style='plain', axis='x')
        self.ax.xaxis.get_offset_text().set_visible(False)
        self.ax.xaxis.set_major_locator(MaxNLocator(integer=True))
        self.ax.ticklabel_format(style='plain', axis='y')
        self.ax.yaxis.get_offset_text().set_visible(False)

        if self.xy_data:
            pts = np.array(self.xy_data, dtype=float)
            pts = pts[np.all(np.isfinite(pts), axis=1)]
            if pts.ndim == 2 and pts.shape[1] == 2 and pts.shape[0] > 0:
                try:
                    from scipy.spatial import KDTree
                    self.kdtree = KDTree(pts)
                except ValueError as e:
                    self.kdtree = None
                    messagebox.showerror("Plot Error", f"KDTree construction failed: {e}")
                    return
            else:
                self.kdtree = None
        else:
            self.kdtree = None

        self.ax.set_title(self.chart_title_entry.get(), pad=15)
        
        for ann_data in saved_manual_annotations:
            ann = self.ax.annotate(ann_data['text'], xy=ann_data['xy'], xytext=ann_data['xytext'],
                                textcoords="offset points",
                                bbox=dict(boxstyle="round", fc="w"),
                                arrowprops=dict(arrowstyle="->"),
                                picker=True)
            self.manual_annotations.append(ann)

        # Adatta automaticamente i limiti del grafico se ci sono dati
        if self.xy_data:
            # Estrai tutti i valori x e y dai dati plottati
            x_values = [x for x, y in self.xy_data]
            y_values = [y for x, y in self.xy_data]
            
            if x_values and y_values:
                # Calcola i limiti con un margine del 5%
                x_min, x_max = min(x_values), max(x_values)
                y_min, y_max = min(y_values), max(y_values)
                
                x_margin = (x_max - x_min) * 0.05
                y_margin = (y_max - y_min) * 0.05
                
                # Imposta i nuovi limiti con margine
                self.ax.set_xlim(x_min - x_margin, x_max + x_margin)
                self.ax.set_ylim(y_min - y_margin, y_max + y_margin)

        if self.initialization_plot == False:
            self.canvas.draw()
            self.initialization_plot = True
        else:
            self.canvas.draw()
            self.firstplot = True


        

def rapid_analysis(main_frame, checkbox_align, start_time1_entry, start_time2_entry, checkbox_event):
    if os.path.isfile('output_file.txt'):
        with open("output_file.txt", "r") as file:
            line = file.readline().strip()
        print(line)
        if os.path.isfile(line):
            os.remove(line)
    if os.path.isfile('output_file.txt'):
        os.remove('output_file.txt')

    if os.path.isfile('output0.csv'):
        df1 = pd.read_csv('output0.csv')
    else:
        messagebox.showerror("Error", "No data found")
        return
    if os.path.isfile('output1.csv'):
        df2 = pd.read_csv('output1.csv')
    if checkbox_align:
        ref_time1 = pd.to_datetime(start_time1_entry.get(), format='%H:%M:%S')
        column_date1 = df1.columns[0]
        df1 = apply_formulas_to_column(df1, ref_time1, column_date1)
        if os.path.isfile('output1.csv'):
            ref_time2 = pd.to_datetime(start_time2_entry.get(), format='%H:%M:%S')
            column_date2 = df2.columns[0]
            df2 = apply_formulas_to_column(df2, ref_time2, column_date2)
            df1, df2, delta_sec = align_dataframes(df1, df2, column_date1, column_date2)
            df1, df2 = process_event(checkbox_event, df1, df2=df2, options_file='options_event.txt', sec=delta_sec)
            launch_interactive_plot(main_frame, df1, df2=df2)
        else:
            df1 = process_event(checkbox_event, df1, df2=None, options_file='options_event.txt', sec=None)
            launch_interactive_plot(main_frame, df1, df2=None)
    else:
        ref_time1 = pd.to_datetime(start_time1_entry.get(), format='%H:%M:%S')
        column_date1 = df1.columns[0]
        df1 = apply_formulas_to_column(df1, ref_time1, column_date1)
        launch_interactive_plot(main_frame, df1, df2=None)
        
def launch_interactive_plot(parent, df1, df2=None):
    InteractivePlotApp(parent, df1, df2)
